#!/usr/bin/env python3
"""
App de Monitoreo de Eficiencia de Mantenimiento
Lee directamente desde SQLite - todos los datos son dinámicos
"""

from flask import Flask, jsonify, request, send_from_directory
import sqlite3
import os
import json

app = Flask(__name__, static_folder='static')
DB_PATH = os.environ.get('DB_PATH', 'mantenimiento.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ── PORTADA ────────────────────────────────────────────────────────────────────
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'portada.html'), 'r', encoding='utf-8') as _f:
    PORTADA_HTML = _f.read()

# ── HTML Principal ─────────────────────────────────────────────────────────────
HTML = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Monitor de Mantenimiento</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');

  :root {
    --bg:        #0f1117;
    --surface:   #171b26;
    --surface2:  #1e2332;
    --border:    #2a3045;
    --accent:    #4f8ef7;
    --accent2:   #f7a94f;
    --accent3:   #4ff7b0;
    --danger:    #f74f4f;
    --text:      #e8ecf4;
    --text2:     #8892a4;
    --mono:      'IBM Plex Mono', monospace;
    --sans:      'IBM Plex Sans', sans-serif;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }

  body {
    background: var(--bg);
    color: var(--text);
    font-family: var(--sans);
    font-size: 14px;
    min-height: 100vh;
  }

  /* HEADER */
  .header {
    background: var(--surface);
    border-bottom: 1px solid var(--border);
    padding: 0 32px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 56px;
    position: sticky;
    top: 0;
    z-index: 100;
  }
  .header-logo {
    font-family: var(--mono);
    font-size: 13px;
    font-weight: 600;
    color: var(--accent);
    letter-spacing: 0.08em;
    text-transform: uppercase;
  }
  .header-logo span { color: var(--text2); font-weight: 400; }

  /* NAV TABS */
  .nav {
    display: flex;
    gap: 4px;
    padding: 0 32px;
    background: var(--surface);
    border-bottom: 1px solid var(--border);
  }
  .nav-tab {
    padding: 10px 16px;
    font-size: 12px;
    font-weight: 500;
    color: var(--text2);
    cursor: pointer;
    border-bottom: 2px solid transparent;
    transition: all 0.15s;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    font-family: var(--mono);
  }
  .nav-tab:hover { color: var(--text); }
  .nav-tab.active { color: var(--accent); border-bottom-color: var(--accent); }

  /* FILTERS */
  .filters {
    padding: 16px 32px;
    background: var(--surface2);
    border-bottom: 1px solid var(--border);
    display: flex;
    gap: 12px;
    flex-wrap: wrap;
    align-items: center;
  }
  .filter-label {
    font-size: 11px;
    font-family: var(--mono);
    color: var(--text2);
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }
  select, input[type=date] {
    background: var(--surface);
    border: 1px solid var(--border);
    color: var(--text);
    padding: 6px 10px;
    font-size: 12px;
    font-family: var(--sans);
    border-radius: 4px;
    cursor: pointer;
    outline: none;
  }
  select:focus, input:focus { border-color: var(--accent); }
  .btn-apply {
    background: var(--accent);
    color: #fff;
    border: none;
    padding: 7px 16px;
    font-size: 12px;
    font-family: var(--mono);
    font-weight: 600;
    border-radius: 4px;
    cursor: pointer;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    transition: opacity 0.15s;
  }
  .btn-apply:hover { opacity: 0.85; }

  /* MAIN */
  .main { padding: 24px 32px; }

  /* VIEWS */
  .view { display: none; }
  .view.active { display: block; }

  /* KPI ROW */
  .kpi-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 12px;
    margin-bottom: 24px;
  }
  .kpi {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 16px 20px;
    position: relative;
    overflow: hidden;
  }
  .kpi::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: var(--kpi-color, var(--accent));
  }
  .kpi-label {
    font-size: 10px;
    font-family: var(--mono);
    color: var(--text2);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 8px;
  }
  .kpi-value {
    font-size: 28px;
    font-family: var(--mono);
    font-weight: 600;
    color: var(--kpi-color, var(--accent));
    line-height: 1;
  }
  .kpi-sub {
    font-size: 11px;
    color: var(--text2);
    margin-top: 4px;
  }

  /* GRID 2 COL */
  .grid-2 {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 16px;
    margin-bottom: 16px;
  }
  .grid-3 {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 16px;
    margin-bottom: 16px;
  }
  @media (max-width: 900px) {
    .grid-2, .grid-3 { grid-template-columns: 1fr; }
  }

  /* CARD */
  .card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 6px;
    overflow: hidden;
  }
  .card-header {
    padding: 12px 16px;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: center;
    justify-content: space-between;
  }
  .card-title {
    font-size: 11px;
    font-family: var(--mono);
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    color: var(--text2);
  }
  .card-body { padding: 16px; }
  .chart-wrap { position: relative; height: 220px; }

  /* TABLE */
  .tbl { width: 100%; border-collapse: collapse; }
  .tbl th {
    font-size: 10px;
    font-family: var(--mono);
    font-weight: 600;
    color: var(--text2);
    text-transform: uppercase;
    letter-spacing: 0.06em;
    padding: 8px 12px;
    text-align: left;
    border-bottom: 1px solid var(--border);
  }
  .tbl td {
    padding: 9px 12px;
    font-size: 12px;
    border-bottom: 1px solid var(--border);
    color: var(--text);
  }
  .tbl tr:last-child td { border-bottom: none; }
  .tbl tr:hover td { background: var(--surface2); }
  .tbl td.num { font-family: var(--mono); text-align: right; }
  .tbl td.mono { font-family: var(--mono); }

  /* BADGES */
  .badge {
    display: inline-block;
    padding: 2px 8px;
    border-radius: 3px;
    font-size: 10px;
    font-family: var(--mono);
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.04em;
  }
  .badge-green  { background: rgba(79,247,176,0.12); color: var(--accent3); }
  .badge-yellow { background: rgba(247,169,79,0.12); color: var(--accent2); }
  .badge-red    { background: rgba(247,79,79,0.12);  color: var(--danger); }
  .badge-blue   { background: rgba(79,142,247,0.12); color: var(--accent); }

  /* BAR INLINE */
  .bar-wrap { display: flex; align-items: center; gap: 8px; }
  .bar-track {
    flex: 1; height: 4px;
    background: var(--border);
    border-radius: 2px;
    overflow: hidden;
  }
  .bar-fill { height: 100%; border-radius: 2px; transition: width 0.4s; }

  /* LOADING */
  .loading {
    display: flex; align-items: center; justify-content: center;
    height: 200px; color: var(--text2);
    font-family: var(--mono); font-size: 12px;
  }
  .loading::before {
    content: ''; display: inline-block;
    width: 16px; height: 16px;
    border: 2px solid var(--border);
    border-top-color: var(--accent);
    border-radius: 50%;
    animation: spin 0.6s linear infinite;
    margin-right: 10px;
  }
  @keyframes spin { to { transform: rotate(360deg); } }

  /* ALERT */
  .alert {
    background: rgba(247,79,79,0.08);
    border: 1px solid rgba(247,79,79,0.25);
    border-radius: 6px;
    padding: 12px 16px;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 10px;
    font-size: 12px;
  }
  .alert-icon { color: var(--danger); font-size: 16px; }

  /* DB INFO */
  .db-info {
    font-family: var(--mono);
    font-size: 11px;
    color: var(--text2);
    display: flex;
    align-items: center;
    gap: 6px;
  }
  .db-dot {
    width: 6px; height: 6px;
    border-radius: 50%;
    background: var(--accent3);
    animation: pulse 2s infinite;
  }
  @keyframes pulse {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.4; }
  }
</style>
</head>
<body>

<div class="header">
  <a href="/" style="display:flex;align-items:center;gap:6px;color:var(--text2);font-family:var(--mono);font-size:11px;text-decoration:none;padding:4px 10px;border:1px solid var(--border);border-radius:4px;margin-right:12px" onmouseover="this.style.color='var(--accent)'" onmouseout="this.style.color='var(--text2)'">&#8592; Empresas</a>
  <div class="header-logo">Monitor<span> / </span>Mantenimiento</div>
  <div class="db-info">
    <div class="db-dot"></div>
    <span id="db-status">cargando...</span>
  </div>
</div>

<div class="nav">
  <div class="nav-tab active" onclick="switchView('resumen')">Resumen</div>
  <div class="nav-tab" onclick="switchView('proveedores')">Proveedores</div>
  <div class="nav-tab" onclick="switchView('tecnicos')">Técnicos</div>
  <div class="nav-tab" onclick="switchView('activos')">Activos</div>
  <div class="nav-tab" onclick="switchView('repuestos')">Repuestos</div>
  <div class="nav-tab" onclick="switchView('locales')">Locales</div>
  <div class="nav-tab" onclick="switchView('proformas')" style="color:var(--accent3)">&#9650; Proformas</div>
  <div class="nav-tab" onclick="switchView('carga')" style="margin-left:auto;color:var(--accent2)">&#8679; Carga de datos</div>
</div>

<div class="filters">
  <span class="filter-label">Período</span>
  <input type="date" id="f-desde" value="2024-01-01">
  <input type="date" id="f-hasta" value="2025-03-15">
  <span class="filter-label">Local</span>
  <select id="f-local"><option value="">Todos</option></select>
  <span class="filter-label">Familia</span>
  <select id="f-familia"><option value="">Todas</option></select>
  <span class="filter-label">Tipo</span>
  <select id="f-tipo">
    <option value="">Todos</option>
    <option value="preventivo">Preventivo</option>
    <option value="correctivo">Correctivo</option>
    <option value="predictivo">Predictivo</option>
  </select>
  <button class="btn-apply" onclick="loadAll()">↻ Actualizar</button>
</div>

<div class="main">

  <!-- RESUMEN -->
  <div class="view active" id="view-resumen">
    <div class="kpi-row" id="kpi-row">
      <div class="loading">Cargando indicadores...</div>
    </div>
    <div id="alertas-section"></div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Distribución — Tipo de mantenimiento</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-tipos"></canvas></div></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Costo por familia de activos</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-familias"></canvas></div></div>
      </div>
    </div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Evolución mensual de intervenciones</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-mensual"></canvas></div></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Interno vs Externo</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-ejecutor"></canvas></div></div>
      </div>
    </div>
  </div>

  <!-- PROVEEDORES -->
  <div class="view" id="view-proveedores">
    <div class="kpi-row" id="kpi-prov"></div>
    <div class="card" style="margin-bottom:16px">
      <div class="card-header"><span class="card-title">Análisis de proveedores externos</span></div>
      <div class="card-body">
        <table class="tbl" id="tbl-proveedores">
          <thead><tr>
            <th>Proveedor</th>
            <th>Contrato</th>
            <th>Intervenciones</th>
            <th>Hs respuesta prom</th>
            <th>Hs trabajo prom</th>
            <th>Costo prom</th>
            <th>Costo total</th>
            <th>Reincidencias</th>
            <th>Rendimiento</th>
          </tr></thead>
          <tbody id="tbody-proveedores"><tr><td colspan="9"><div class="loading">Cargando...</div></td></tr></tbody>
        </table>
      </div>
    </div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Tiempo de respuesta por proveedor (hs)</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-prov-resp"></canvas></div></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Costo promedio por intervención</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-prov-costo"></canvas></div></div>
      </div>
    </div>
  </div>

  <!-- TÉCNICOS -->
  <div class="view" id="view-tecnicos">
    <div class="kpi-row" id="kpi-tec"></div>
    <div class="card" style="margin-bottom:16px">
      <div class="card-header"><span class="card-title">Eficiencia de técnicos internos</span></div>
      <div class="card-body">
        <table class="tbl">
          <thead><tr>
            <th>Técnico</th>
            <th>Nivel</th>
            <th>Intervenciones</th>
            <th>Hs respuesta prom</th>
            <th>Hs trabajo prom</th>
            <th>Reincidencias</th>
            <th>2da visita</th>
            <th>Costo total</th>
            <th>Eficiencia</th>
          </tr></thead>
          <tbody id="tbody-tecnicos"><tr><td colspan="9"><div class="loading">Cargando...</div></td></tr></tbody>
        </table>
      </div>
    </div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Intervenciones por técnico</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-tec-interv"></canvas></div></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Reincidencias por técnico</th></span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-tec-reinc"></canvas></div></div>
      </div>
    </div>
  </div>

  <!-- ACTIVOS -->
  <div class="view" id="view-activos">
    <div class="kpi-row" id="kpi-act"></div>
    <div class="card" style="margin-bottom:16px">
      <div class="card-header"><span class="card-title">Top activos por costo acumulado</span></div>
      <div class="card-body">
        <table class="tbl">
          <thead><tr>
            <th>Código</th>
            <th>Activo</th>
            <th>Local</th>
            <th>Familia</th>
            <th>Estado</th>
            <th>Intervenciones</th>
            <th>Días paralizado</th>
            <th>Costo acumulado</th>
            <th>Costo reemplazo</th>
            <th>Señal</th>
          </tr></thead>
          <tbody id="tbody-activos"><tr><td colspan="10"><div class="loading">Cargando...</div></td></tr></tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- REPUESTOS -->
  <div class="view" id="view-repuestos">
    <div class="kpi-row" id="kpi-rep"></div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Repuestos más utilizados</span></div>
        <div class="card-body">
          <table class="tbl">
            <thead><tr><th>Repuesto</th><th>Usos</th><th>Precio prom</th><th>Precio max</th><th>Variación</th></tr></thead>
            <tbody id="tbody-repuestos"><tr><td colspan="5"><div class="loading">Cargando...</div></td></tr></tbody>
          </table>
        </div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Variación de precio por repuesto</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-rep-precio"></canvas></div></div>
      </div>
    </div>
  </div>

  <!-- LOCALES -->
  <div class="view" id="view-locales">
    <div class="kpi-row" id="kpi-loc"></div>
    <div class="card" style="margin-bottom:16px">
      <div class="card-header"><span class="card-title">Análisis por local / sucursal</span></div>
      <div class="card-body">
        <table class="tbl">
          <thead><tr>
            <th>Local</th>
            <th>Ciudad</th>
            <th>Intervenciones</th>
            <th>% Correctivo</th>
            <th>Hs respuesta prom</th>
            <th>Costo total</th>
            <th>Costo / m²</th>
            <th>Días paralizado</th>
          </tr></thead>
          <tbody id="tbody-locales"><tr><td colspan="8"><div class="loading">Cargando...</div></td></tr></tbody>
        </table>
      </div>
    </div>
    <div class="grid-2">
      <div class="card">
        <div class="card-header"><span class="card-title">Costo total por local</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-loc-costo"></canvas></div></div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">% Correctivo por local</span></div>
        <div class="card-body"><div class="chart-wrap"><canvas id="chart-loc-correctivo"></canvas></div></div>
      </div>
    </div>
  </div>

  <!-- CARGA DE DATOS -->
  <div class="view" id="view-carga">
    <div class="kpi-row">
      <div class="kpi" style="--kpi-color:var(--accent2)">
        <div class="kpi-label">Módulo de carga</div>
        <div class="kpi-value" style="font-size:18px">Excel → Base</div>
        <div class="kpi-sub">Subí el Excel con los datos y se cargan automáticamente</div>
      </div>
    </div>

    <div class="grid-2" style="margin-bottom:16px">
      <!-- Panel subida -->
      <div class="card">
        <div class="card-header">
          <span class="card-title">Subir archivo Excel</span>
        </div>
        <div class="card-body" style="display:flex;flex-direction:column;gap:16px">
          <div style="font-size:12px;color:var(--text2);line-height:1.6">
            Usá la plantilla oficial con las hojas:<br>
            <strong style="color:var(--text)">1. Activos · 2. Intervenciones · 3. Proveedores · 4. Técnicos · 5. Locales</strong><br>
            Podés subir solo las hojas que tengas completadas.
          </div>
          <div id="drop-zone" style="border:2px dashed var(--border);border-radius:8px;padding:32px;text-align:center;cursor:pointer;transition:all 0.2s"
               onclick="document.getElementById('file-input').click()"
               ondragover="event.preventDefault();this.style.borderColor='var(--accent)'"
               ondragleave="this.style.borderColor='var(--border)'"
               ondrop="handleDrop(event)">
            <div style="font-size:28px;margin-bottom:8px">📂</div>
            <div style="font-size:13px;color:var(--text2)">Hacer clic o arrastrar el archivo Excel aquí</div>
            <div style="font-size:11px;color:var(--text2);margin-top:4px;font-family:var(--mono)">.xlsx únicamente</div>
          </div>
          <input type="file" id="file-input" accept=".xlsx" style="display:none" onchange="handleFile(this.files[0])">
          <div id="file-name" style="font-family:var(--mono);font-size:11px;color:var(--accent);display:none"></div>
          <button id="btn-cargar" onclick="subirArchivo()" disabled
            style="background:var(--accent);color:#fff;border:none;padding:10px 20px;border-radius:6px;font-size:13px;font-family:var(--mono);font-weight:600;cursor:pointer;opacity:0.4;transition:opacity 0.2s">
            ↑ Cargar en la base
          </button>
        </div>
      </div>

      <!-- Estado de carga -->
      <div class="card">
        <div class="card-header">
          <span class="card-title">Resultado de la carga</span>
        </div>
        <div class="card-body">
          <div id="carga-resultado" style="font-family:var(--mono);font-size:12px;color:var(--text2);min-height:200px">
            Esperando archivo...
          </div>
        </div>
      </div>
    </div>

    <!-- Resumen de la base actual -->
    <div class="card">
      <div class="card-header">
        <span class="card-title">Estado actual de la base</span>
        <button onclick="limpiarBase()" style="background:rgba(247,79,79,0.15);border:1px solid rgba(247,79,79,0.4);color:var(--danger);padding:4px 12px;border-radius:4px;font-size:11px;font-family:var(--mono);cursor:pointer">
          &#x1F5D1; Limpiar datos cargados
        </button>
      </div>
      <div class="card-body">
        <table class="tbl" id="tbl-estado-base">
          <thead><tr>
            <th>Tabla</th><th>Registros actuales</th><th>Última carga</th>
          </tr></thead>
          <tbody id="tbody-estado-base">
            <tr><td colspan="3"><div class="loading">Cargando...</div></td></tr>
          </tbody>
        </table>
      </div>
    </div>
  </div>


  <!-- PROFORMAS -->
  <div class="view" id="view-proformas">
    <div class="kpi-row">
      <div class="kpi" style="--kpi-color:var(--accent3)">
        <div class="kpi-label">Modulo proformas</div>
        <div class="kpi-value" style="font-size:18px">CSV/Excel</div>
        <div class="kpi-sub">Subí el archivo y detectamos patrones</div>
      </div>
    </div>
    <div class="grid-2" style="margin-bottom:16px">
      <div class="card">
        <div class="card-header"><span class="card-title">Subir archivo de proformas</span></div>
        <div class="card-body" style="display:flex;flex-direction:column;gap:14px">
          <div style="font-size:12px;color:var(--text2)">Subí el CSV o Excel exportado del sistema. La app detecta automaticamente proveedores, locales y riesgo financiero.</div>
          <div id="pf-drop" style="border:2px dashed var(--border);border-radius:8px;padding:28px;text-align:center;cursor:pointer" onclick="document.getElementById('pf-fi').click()">
            <div style="font-size:24px;margin-bottom:8px">&#128203;</div>
            <div style="font-size:13px;color:var(--text2)">Clic o arrastrar archivo aqui</div>
            <div style="font-size:11px;color:var(--text2);margin-top:4px;font-family:var(--mono)">.csv o .xlsx</div>
          </div>
          <input type="file" id="pf-fi" accept=".csv,.xlsx" style="display:none" onchange="pfSelec(this.files[0])">
          <div id="pf-fn" style="font-family:var(--mono);font-size:11px;color:var(--accent3);display:none"></div>
          <button id="pf-btn" onclick="pfAnalizar()" disabled style="background:var(--accent3);color:#000;border:none;padding:10px 20px;border-radius:6px;font-size:13px;font-family:var(--mono);font-weight:600;cursor:pointer;opacity:0.4">
            Analizar proformas
          </button>
        </div>
      </div>
      <div class="card">
        <div class="card-header"><span class="card-title">Resumen</span></div>
        <div class="card-body">
          <div id="pf-res" style="font-family:var(--mono);font-size:12px;color:var(--text2);min-height:180px">Esperando archivo...</div>
        </div>
      </div>
    </div>
    <div id="pf-kpis" class="kpi-row" style="display:none"></div>
    <div id="pf-tablas" style="display:none">
      <div class="grid-2" style="margin-bottom:16px">
        <div class="card">
          <div class="card-header"><span class="card-title">Sin factura por proveedor</span></div>
          <div class="card-body">
            <table class="tbl"><thead><tr><th>Proveedor</th><th>Sin factura</th><th>Total</th><th>% riesgo</th></tr></thead>
            <tbody id="pf-t1"></tbody></table>
          </div>
        </div>
        <div class="card">
          <div class="card-header"><span class="card-title">Concentracion por proveedor</span></div>
          <div class="card-body">
            <table class="tbl"><thead><tr><th>Proveedor</th><th>N</th><th>Monto total</th><th>Ticket prom</th></tr></thead>
            <tbody id="pf-t2"></tbody></table>
          </div>
        </div>
      </div>
      <div class="grid-2" style="margin-bottom:16px">
        <div class="card">
          <div class="card-header"><span class="card-title">Gasto por local</span></div>
          <div class="card-body">
            <table class="tbl"><thead><tr><th>Local</th><th>N</th><th>Monto total</th><th>Ticket prom</th></tr></thead>
            <tbody id="pf-t3"></tbody></table>
          </div>
        </div>
        <div class="card">
          <div class="card-header"><span class="card-title">Por responsable</span></div>
          <div class="card-body">
            <table class="tbl"><thead><tr><th>Responsable</th><th>N</th><th>Monto total</th><th>%</th></tr></thead>
            <tbody id="pf-t4"></tbody></table>
          </div>
        </div>
      </div>
      <div class="card" style="margin-bottom:16px">
        <div class="card-header"><span class="card-title">Tipos de trabajo mas frecuentes</span></div>
        <div class="card-body">
          <table class="tbl"><thead><tr><th>Tipo</th><th>N</th><th>Monto total</th><th>Ticket prom</th></tr></thead>
          <tbody id="pf-t5"></tbody></table>
        </div>
      </div>
    </div>
  </div>

</div><!-- /main -->

<script>
// ── Estado global ─────────────────────────────────────────────────────────────
let charts = {};
let currentView = 'resumen';

// ── Utilidades ────────────────────────────────────────────────────────────────
const $ = id => document.getElementById(id);
const fmt = n => n == null ? '-' : Number(n).toLocaleString('es-AR', {maximumFractionDigits: 0});
const fmtDec = (n, d=1) => n == null ? '-' : Number(n).toFixed(d);
const fmtPeso = n => n == null ? '-' : '$' + fmt(n);

function getFilters() {
  return {
    desde:   $('f-desde').value,
    hasta:   $('f-hasta').value,
    local:   $('f-local').value,
    familia: $('f-familia').value,
    tipo:    $('f-tipo').value,
  };
}

function params(extra = {}) {
  const f = {...getFilters(), ...extra};
  return '?' + Object.entries(f).filter(([,v]) => v).map(([k,v]) => `${k}=${encodeURIComponent(v)}`).join('&');
}

async function api(endpoint) {
  const r = await fetch('/api/' + endpoint);
  return r.json();
}

function destroyChart(id) {
  if (charts[id]) { charts[id].destroy(); delete charts[id]; }
}

function makeChart(id, type, labels, datasets, opts = {}) {
  destroyChart(id);
  const ctx = $(id);
  if (!ctx) return;
  charts[id] = new Chart(ctx, {
    type,
    data: { labels, datasets },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      plugins: {
        legend: { labels: { color: '#8892a4', font: { size: 11, family: "'IBM Plex Sans'" } } },
        tooltip: { backgroundColor: '#1e2332', borderColor: '#2a3045', borderWidth: 1,
                   titleColor: '#e8ecf4', bodyColor: '#8892a4', padding: 10 }
      },
      scales: type !== 'pie' && type !== 'doughnut' ? {
        x: { ticks: { color: '#8892a4', font: { size: 10 } }, grid: { color: '#2a3045' } },
        y: { ticks: { color: '#8892a4', font: { size: 10 } }, grid: { color: '#2a3045' } }
      } : {},
      ...opts
    }
  });
}

function badgeEstado(e) {
  const m = {bueno:'badge-green', regular:'badge-yellow', critico:'badge-red', fuera_de_servicio:'badge-red'};
  return `<span class="badge ${m[e]||'badge-blue'}">${e||'-'}</span>`;
}

function badgeContrato(c) {
  const m = {anual:'badge-green', por_equipo:'badge-blue', puntual:'badge-yellow', sin_contrato:'badge-red'};
  const l = {anual:'Anual', por_equipo:'Por equipo', puntual:'Puntual', sin_contrato:'Sin contrato'};
  return `<span class="badge ${m[c]||'badge-blue'}">${l[c]||c}</span>`;
}

function renderBar(val, max, color='var(--accent)') {
  const pct = Math.min(100, (val/max)*100);
  return `<div class="bar-wrap">
    <span style="font-family:var(--mono);font-size:11px;min-width:40px">${fmtDec(val)}</span>
    <div class="bar-track"><div class="bar-fill" style="width:${pct}%;background:${color}"></div></div>
  </div>`;
}

function renderKpis(containerId, kpis) {
  $(containerId).innerHTML = kpis.map(k => `
    <div class="kpi" style="--kpi-color:${k.color||'var(--accent)'}">
      <div class="kpi-label">${k.label}</div>
      <div class="kpi-value">${k.value}</div>
      <div class="kpi-sub">${k.sub||''}</div>
    </div>
  `).join('');
}

// ── SWITCH VIEW ───────────────────────────────────────────────────────────────
function switchView(name) {
  currentView = name;
  document.querySelectorAll('.view').forEach(v => v.classList.remove('active'));
  document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
  $('view-'+name).classList.add('active');
  event.target.classList.add('active');
  loadView(name);
}

// ── LOAD ALL ──────────────────────────────────────────────────────────────────
async function loadAll() {
  await loadFilters();
  await loadView(currentView);
}

async function loadFilters() {
  const [locales, familias] = await Promise.all([api('filtros/locales'), api('filtros/familias')]);
  const sel = id => {
    const prev = $(id).value;
    return prev;
  };
  
  const fLocal = sel('f-local');
  $('f-local').innerHTML = '<option value="">Todos</option>' +
    locales.map(l => `<option value="${l.id}" ${l.id==fLocal?'selected':''}>${l.nombre}</option>`).join('');
  
  const fFam = sel('f-familia');
  $('f-familia').innerHTML = '<option value="">Todas</option>' +
    familias.map(f => `<option value="${f.id}" ${f.id==fFam?'selected':''}>${f.nombre}</option>`).join('');
}

async function loadView(name) {
  switch(name) {
    case 'resumen':    await loadResumen(); break;
    case 'proveedores': await loadProveedores(); break;
    case 'tecnicos':   await loadTecnicos(); break;
    case 'activos':    await loadActivos(); break;
    case 'repuestos':  await loadRepuestos(); break;
    case 'locales':    await loadLocales(); break;
    case 'proformas':   break;
    case 'carga':      await cargarEstadoBase(); break;
  }
}

// ── RESUMEN ───────────────────────────────────────────────────────────────────
async function loadResumen() {
  const [kpis, tipos, familias, mensual, ejecutor, alertas, info] = await Promise.all([
    api('resumen/kpis'+params()),
    api('resumen/tipos'+params()),
    api('resumen/familias'+params()),
    api('resumen/mensual'+params()),
    api('resumen/ejecutor'+params()),
    api('resumen/alertas'+params()),
    api('info'),
  ]);

  $('db-status').textContent = `${info.total_intervenciones} intervenciones · ${info.total_activos} activos · ${info.db_file}`;

  renderKpis('kpi-row', [
    { label: 'Total intervenciones', value: fmt(kpis.total_intervenciones), sub: 'en el período', color: 'var(--accent)' },
    { label: 'Costo total', value: fmtPeso(kpis.costo_total), sub: 'mantenimiento', color: 'var(--accent2)' },
    { label: '% Correctivo', value: fmtDec(kpis.pct_correctivo)+'%', sub: 'del total', color: kpis.pct_correctivo > 60 ? 'var(--danger)' : 'var(--accent3)' },
    { label: 'Reincidencias', value: fmt(kpis.total_reincidencias), sub: 'misma falla repetida', color: 'var(--danger)' },
    { label: 'Días paralizado', value: fmtDec(kpis.dias_paralizado), sub: 'equipos fuera de servicio', color: 'var(--accent2)' },
    { label: 'Proveedores activos', value: fmt(kpis.proveedores_activos), sub: 'en el período', color: 'var(--accent)' },
  ]);

  // Alertas
  const alertasHtml = alertas.map(a => `
    <div class="alert">
      <span class="alert-icon">⚠</span>
      <span>${a.mensaje}</span>
    </div>
  `).join('');
  $('alertas-section').innerHTML = alertasHtml;

  // Gráfico tipos
  const coloresTipos = { preventivo: '#4ff7b0', correctivo: '#f74f4f', predictivo: '#4f8ef7' };
  makeChart('chart-tipos', 'doughnut',
    tipos.map(t => t.tipo),
    [{
      data: tipos.map(t => t.cantidad),
      backgroundColor: tipos.map(t => coloresTipos[t.tipo] || '#4f8ef7'),
      borderWidth: 0,
    }]
  );

  // Gráfico familias
  makeChart('chart-familias', 'bar',
    familias.map(f => f.familia.replace(' y ', ' & ')),
    [{
      label: 'Costo total',
      data: familias.map(f => f.costo_total),
      backgroundColor: '#4f8ef780',
      borderColor: '#4f8ef7',
      borderWidth: 1,
    }]
  );

  // Gráfico mensual
  makeChart('chart-mensual', 'line',
    mensual.map(m => m.mes),
    [{
      label: 'Intervenciones',
      data: mensual.map(m => m.cantidad),
      borderColor: '#4f8ef7',
      backgroundColor: '#4f8ef720',
      fill: true,
      tension: 0.3,
    }]
  );

  // Ejecutor
  makeChart('chart-ejecutor', 'doughnut',
    ejecutor.map(e => e.tipo_ejecutor === 'interno' ? 'Interno' : 'Externo'),
    [{
      data: ejecutor.map(e => e.cantidad),
      backgroundColor: ['#4ff7b0', '#f7a94f'],
      borderWidth: 0,
    }]
  );
}

// ── PROVEEDORES ───────────────────────────────────────────────────────────────
async function loadProveedores() {
  const data = await api('proveedores'+params());
  const maxResp = Math.max(...data.map(d => d.hs_respuesta||0));
  const maxCosto = Math.max(...data.map(d => d.costo_prom||0));

  renderKpis('kpi-prov', [
    { label: 'Proveedores analizados', value: data.length, color: 'var(--accent)' },
    { label: 'Más rápido', value: data.sort((a,b)=>a.hs_respuesta-b.hs_respuesta)[0]?.nombre?.split(' ')[0]||'-', sub: fmtDec(data[0]?.hs_respuesta)+' hs resp', color: 'var(--accent3)' },
    { label: 'Más costoso', value: data.sort((a,b)=>b.costo_prom-a.costo_prom)[0]?.nombre?.split(' ')[0]||'-', sub: fmtPeso(data[0]?.costo_prom)+' prom', color: 'var(--danger)' },
  ]);
  data.sort((a,b) => a.hs_respuesta - b.hs_respuesta);

  $('tbody-proveedores').innerHTML = data.map(d => {
    const reinc_pct = d.intervenciones > 0 ? (d.reincidencias/d.intervenciones*100) : 0;
    const rendimiento = d.hs_respuesta < 8 && reinc_pct < 10 ? 'badge-green' :
                        d.hs_respuesta > 15 || reinc_pct > 20 ? 'badge-red' : 'badge-yellow';
    const rendLabel = d.hs_respuesta < 8 && reinc_pct < 10 ? 'Bueno' :
                      d.hs_respuesta > 15 || reinc_pct > 20 ? 'Deficiente' : 'Regular';
    return `<tr>
      <td><strong>${d.nombre}</strong></td>
      <td>${badgeContrato(d.tipo_contrato)}</td>
      <td class="num">${fmt(d.intervenciones)}</td>
      <td>${renderBar(d.hs_respuesta, maxResp, d.hs_respuesta > 12 ? 'var(--danger)' : 'var(--accent3)')}</td>
      <td class="num">${fmtDec(d.hs_trabajo)} hs</td>
      <td class="num">${fmtPeso(d.costo_prom)}</td>
      <td class="num">${fmtPeso(d.costo_total)}</td>
      <td class="num">${fmt(d.reincidencias)} <span style="color:var(--text2);font-size:10px">(${fmtDec(reinc_pct)}%)</span></td>
      <td><span class="badge ${rendimiento}">${rendLabel}</span></td>
    </tr>`;
  }).join('');

  const sorted_resp = [...data].sort((a,b) => b.hs_respuesta - a.hs_respuesta);
  makeChart('chart-prov-resp', 'bar',
    sorted_resp.map(d => d.nombre.split(' ')[0]),
    [{
      label: 'Hs respuesta prom',
      data: sorted_resp.map(d => d.hs_respuesta),
      backgroundColor: sorted_resp.map(d => d.hs_respuesta > 12 ? '#f74f4f80' : '#4ff7b080'),
      borderColor: sorted_resp.map(d => d.hs_respuesta > 12 ? '#f74f4f' : '#4ff7b0'),
      borderWidth: 1,
    }]
  );

  const sorted_costo = [...data].sort((a,b) => b.costo_prom - a.costo_prom);
  makeChart('chart-prov-costo', 'bar',
    sorted_costo.map(d => d.nombre.split(' ')[0]),
    [{
      label: 'Costo promedio',
      data: sorted_costo.map(d => d.costo_prom),
      backgroundColor: '#f7a94f80',
      borderColor: '#f7a94f',
      borderWidth: 1,
    }]
  );
}

// ── TÉCNICOS ──────────────────────────────────────────────────────────────────
async function loadTecnicos() {
  const data = await api('tecnicos'+params());

  renderKpis('kpi-tec', [
    { label: 'Técnicos activos', value: data.length, color: 'var(--accent)' },
    { label: 'Más eficiente', value: data.sort((a,b)=>a.hs_trabajo-b.hs_trabajo)[0]?.nombre?.split(' ')[0]||'-', sub: fmtDec(data[0]?.hs_trabajo)+' hs prom', color: 'var(--accent3)' },
    { label: 'Más reincidencias', value: data.sort((a,b)=>b.reincidencias-a.reincidencias)[0]?.nombre?.split(' ')[0]||'-', sub: fmt(data[0]?.reincidencias)+' casos', color: 'var(--danger)' },
  ]);
  data.sort((a,b) => a.hs_trabajo - b.hs_trabajo);

  const maxReinc = Math.max(...data.map(d => d.reincidencias||0));

  $('tbody-tecnicos').innerHTML = data.map(d => {
    const reinc_pct = d.intervenciones > 0 ? (d.reincidencias/d.intervenciones*100) : 0;
    const efic = reinc_pct < 10 && d.hs_trabajo < 4 ? 'badge-green' :
                 reinc_pct > 25 || d.hs_trabajo > 5 ? 'badge-red' : 'badge-yellow';
    const eficLabel = reinc_pct < 10 && d.hs_trabajo < 4 ? 'Alto' :
                      reinc_pct > 25 || d.hs_trabajo > 5 ? 'Bajo' : 'Medio';
    return `<tr>
      <td><strong>${d.nombre}</strong></td>
      <td><span class="badge badge-blue">${d.nivel}</span></td>
      <td class="num">${fmt(d.intervenciones)}</td>
      <td class="num">${fmtDec(d.hs_respuesta)} hs</td>
      <td class="num">${fmtDec(d.hs_trabajo)} hs</td>
      <td>${renderBar(d.reincidencias, maxReinc, d.reincidencias > 8 ? 'var(--danger)' : 'var(--accent3)')}</td>
      <td class="num">${fmt(d.segunda_visita)}</td>
      <td class="num">${fmtPeso(d.costo_total)}</td>
      <td><span class="badge ${efic}">${eficLabel}</span></td>
    </tr>`;
  }).join('');

  makeChart('chart-tec-interv', 'bar',
    data.map(d => d.nombre.split(' ')[0]),
    [{
      label: 'Intervenciones',
      data: data.map(d => d.intervenciones),
      backgroundColor: '#4f8ef780',
      borderColor: '#4f8ef7',
      borderWidth: 1,
    }]
  );

  makeChart('chart-tec-reinc', 'bar',
    data.map(d => d.nombre.split(' ')[0]),
    [{
      label: 'Reincidencias',
      data: data.map(d => d.reincidencias),
      backgroundColor: data.map(d => d.reincidencias > 8 ? '#f74f4f80' : '#4f8ef780'),
      borderColor: data.map(d => d.reincidencias > 8 ? '#f74f4f' : '#4f8ef7'),
      borderWidth: 1,
    }]
  );
}

// ── ACTIVOS ───────────────────────────────────────────────────────────────────
async function loadActivos() {
  const data = await api('activos'+params());

  const totalCosto = data.reduce((s,d) => s + (d.costo_total||0), 0);
  const totalDias = data.reduce((s,d) => s + (d.dias_paralizado||0), 0);

  renderKpis('kpi-act', [
    { label: 'Activos analizados', value: data.length, color: 'var(--accent)' },
    { label: 'Costo total', value: fmtPeso(totalCosto), color: 'var(--accent2)' },
    { label: 'Días paralizado total', value: fmtDec(totalDias), color: 'var(--danger)' },
    { label: 'Candidatos reemplazo', value: data.filter(d => d.costo_total > d.costo_reemplazo*0.5).length, sub: 'costo > 50% reemplazo', color: 'var(--danger)' },
  ]);

  $('tbody-activos').innerHTML = data.slice(0,30).map(d => {
    const ratio = d.costo_reemplazo > 0 ? (d.costo_total / d.costo_reemplazo * 100) : 0;
    const senal = ratio > 50 ? '<span class="badge badge-red">Reemplazar</span>' :
                  ratio > 25 ? '<span class="badge badge-yellow">Evaluar</span>' :
                  '<span class="badge badge-green">OK</span>';
    return `<tr>
      <td class="mono">${d.codigo}</td>
      <td>${d.descripcion}</td>
      <td>${d.local}</td>
      <td>${d.familia}</td>
      <td>${badgeEstado(d.estado)}</td>
      <td class="num">${fmt(d.intervenciones)}</td>
      <td class="num" style="color:${d.dias_paralizado>5?'var(--danger)':'inherit'}">${fmtDec(d.dias_paralizado)}</td>
      <td class="num">${fmtPeso(d.costo_total)}</td>
      <td class="num">${fmtPeso(d.costo_reemplazo)}</td>
      <td>${senal}</td>
    </tr>`;
  }).join('');
}

// ── REPUESTOS ─────────────────────────────────────────────────────────────────
async function loadRepuestos() {
  const data = await api('repuestos'+params());
  const maxVar = Math.max(...data.map(d => d.variacion_pct||0));

  renderKpis('kpi-rep', [
    { label: 'Repuestos distintos', value: data.length, color: 'var(--accent)' },
    { label: 'Mayor variación precio', value: data.sort((a,b)=>b.variacion_pct-a.variacion_pct)[0]?.nombre?.split(' ')[0]||'-', sub: fmtDec(data[0]?.variacion_pct)+'% var', color: 'var(--danger)' },
  ]);
  data.sort((a,b) => b.usos - a.usos);

  $('tbody-repuestos').innerHTML = data.map(d => {
    return `<tr>
      <td>${d.nombre}</td>
      <td class="num">${fmt(d.usos)}</td>
      <td class="num">${fmtPeso(d.precio_prom)}</td>
      <td class="num">${fmtPeso(d.precio_max)}</td>
      <td>${renderBar(d.variacion_pct, maxVar, d.variacion_pct > 30 ? 'var(--danger)' : 'var(--accent2)')}</td>
    </tr>`;
  }).join('');

  const sorted = [...data].sort((a,b) => b.variacion_pct - a.variacion_pct).slice(0,8);
  makeChart('chart-rep-precio', 'bar',
    sorted.map(d => d.nombre.split(' ').slice(0,2).join(' ')),
    [{
      label: '% variación precio',
      data: sorted.map(d => d.variacion_pct),
      backgroundColor: sorted.map(d => d.variacion_pct > 30 ? '#f74f4f80' : '#f7a94f80'),
      borderColor: sorted.map(d => d.variacion_pct > 30 ? '#f74f4f' : '#f7a94f'),
      borderWidth: 1,
    }]
  );
}

// ── LOCALES ───────────────────────────────────────────────────────────────────
async function loadLocales() {
  const data = await api('locales'+params());
  const maxCosto = Math.max(...data.map(d => d.costo_total||0));

  renderKpis('kpi-loc', [
    { label: 'Locales analizados', value: data.length, color: 'var(--accent)' },
    { label: 'Mayor costo', value: data.sort((a,b)=>b.costo_total-a.costo_total)[0]?.nombre?.split(' ').slice(0,2).join(' ')||'-', color: 'var(--danger)' },
    { label: 'Mayor % correctivo', value: data.sort((a,b)=>b.pct_correctivo-a.pct_correctivo)[0]?.nombre?.split(' ').slice(0,2).join(' ')||'-', sub: fmtDec(data[0]?.pct_correctivo)+'%', color: 'var(--accent2)' },
  ]);
  data.sort((a,b) => b.costo_total - a.costo_total);

  $('tbody-locales').innerHTML = data.map(d => {
    return `<tr>
      <td><strong>${d.nombre}</strong></td>
      <td>${d.ciudad}</td>
      <td class="num">${fmt(d.intervenciones)}</td>
      <td><div class="bar-wrap">
        <span style="font-family:var(--mono);font-size:11px;min-width:36px">${fmtDec(d.pct_correctivo)}%</span>
        <div class="bar-track"><div class="bar-fill" style="width:${d.pct_correctivo}%;background:${d.pct_correctivo>65?'var(--danger)':'var(--accent2)'}"></div></div>
      </div></td>
      <td class="num">${fmtDec(d.hs_respuesta)} hs</td>
      <td class="num">${fmtPeso(d.costo_total)}</td>
      <td class="num">${d.costo_m2 ? fmtPeso(d.costo_m2) : '-'}</td>
      <td class="num" style="color:${d.dias_paralizado>5?'var(--danger)':'inherit'}">${fmtDec(d.dias_paralizado)}</td>
    </tr>`;
  }).join('');

  makeChart('chart-loc-costo', 'bar',
    data.map(d => d.nombre.split(' ').slice(0,2).join(' ')),
    [{
      label: 'Costo total',
      data: data.map(d => d.costo_total),
      backgroundColor: '#4f8ef780',
      borderColor: '#4f8ef7',
      borderWidth: 1,
    }]
  );

  makeChart('chart-loc-correctivo', 'bar',
    data.map(d => d.nombre.split(' ').slice(0,2).join(' ')),
    [{
      label: '% Correctivo',
      data: data.map(d => d.pct_correctivo),
      backgroundColor: data.map(d => d.pct_correctivo > 65 ? '#f74f4f80' : '#4f8ef780'),
      borderColor: data.map(d => d.pct_correctivo > 65 ? '#f74f4f' : '#4f8ef7'),
      borderWidth: 1,
    }]
  );
}

// ── CARGA DE DATOS ───────────────────────────────────────────────────────────
let archivoSeleccionado = null;

function handleFile(file) {
  if (!file) return;
  archivoSeleccionado = file;
  const nameEl = document.getElementById('file-name');
  nameEl.textContent = '📄 ' + file.name + ' (' + (file.size/1024).toFixed(1) + ' KB)';
  nameEl.style.display = 'block';
  const btn = document.getElementById('btn-cargar');
  btn.disabled = false;
  btn.style.opacity = '1';
  document.getElementById('drop-zone').style.borderColor = 'var(--accent)';
}

function handleDrop(e) {
  e.preventDefault();
  document.getElementById('drop-zone').style.borderColor = 'var(--border)';
  const file = e.dataTransfer.files[0];
  if (file && file.name.endsWith('.xlsx')) handleFile(file);
}

async function subirArchivo() {
  if (!archivoSeleccionado) return;
  const btn = document.getElementById('btn-cargar');
  const resultado = document.getElementById('carga-resultado');
  btn.disabled = true;
  btn.textContent = 'Procesando...';
  resultado.innerHTML = '<span style="color:var(--accent2)">⏳ Procesando el archivo...</span>';

  const formData = new FormData();
  formData.append('archivo', archivoSeleccionado);

  try {
    const resp = await fetch('/api/cargar', { method: 'POST', body: formData });
    const data = await resp.json();
    
    if (data.error) {
      resultado.innerHTML = '<span style="color:var(--danger)">✗ Error: ' + data.error + '</span>';
    } else {
      let html = '<span style="color:var(--accent3)">✓ Carga completada</span><br><br>';
      for (const [tabla, info] of Object.entries(data.resultado)) {
        const color = info.insertados > 0 ? 'var(--accent3)' : 'var(--text2)';
        html += '<span style="color:' + color + '">• ' + tabla + ': ' + info.insertados + ' registros nuevos';
        if (info.errores > 0) {
          html += ' <span style="color:var(--danger)">(' + info.errores + ' con errores)</span>';
          if (info.detalles && info.detalles.length > 0) {
            info.detalles.forEach(d => { html += '<br><span style="color:var(--danger);font-size:10px;padding-left:16px">&#8594; ' + d + '</span>'; });
          }
        }
        html += '</span><br>';
      }
      resultado.innerHTML = html;
      await cargarEstadoBase();
      await loadAll();
    }
  } catch(e) {
    resultado.innerHTML = '<span style="color:var(--danger)">✗ Error de conexión: ' + e.message + '</span>';
  }
  
  btn.disabled = false;
  btn.textContent = '↑ Cargar en la base';
}

async function limpiarBase() {
  if (!confirm('¿Seguro? Esto elimina TODOS los datos cargados de intervenciones, activos, proveedores y técnicos. Los datos de demo ficticios se mantienen separados.')) return;
  try {
    const resp = await fetch('/api/limpiar-carga', { method: 'POST' });
    const data = await resp.json();
    if (data.ok) {
      document.getElementById('carga-resultado').innerHTML = '<span style="color:var(--accent3)">✓ Datos limpiados correctamente</span>';
      await cargarEstadoBase();
      await loadAll();
    }
  } catch(e) {
    alert('Error al limpiar: ' + e.message);
  }
}

async function cargarEstadoBase() {
  try {
    const data = await api('estado-base');
    document.getElementById('tbody-estado-base').innerHTML = data.map(r => `
      <tr>
        <td class="mono">${r.tabla}</td>
        <td class="num">${fmt(r.registros)}</td>
        <td style="color:var(--text2);font-size:11px">${r.ultima_carga || 'Sin datos'}</td>
      </tr>
    `).join('');
  } catch(e) {}
}


// PROFORMAS
var pfArchivo = null;
function pfSelec(f) {
  pfArchivo = f;
  var el = document.getElementById('pf-fn');
  el.textContent = f.name + ' (' + Math.round(f.size/1024) + ' KB)';
  el.style.display = 'block';
  var btn = document.getElementById('pf-btn');
  btn.disabled = false;
  btn.style.opacity = '1';
}
async function pfAnalizar() {
  if (!pfArchivo) return;
  var btn = document.getElementById('pf-btn');
  var res = document.getElementById('pf-res');
  btn.textContent = 'Analizando...';
  btn.disabled = true;
  res.innerHTML = '<span style="color:var(--accent3)">Procesando...</span>';
  var fd = new FormData();
  fd.append('archivo', pfArchivo);
  try {
    var resp = await fetch('/api/analizar-proformas', {method:'POST', body:fd});
    var d = await resp.json();
    if (d.error) {
      res.innerHTML = '<span style="color:var(--danger)">Error: ' + d.error + '</span>';
    } else {
      res.innerHTML = '<span style="color:var(--accent3)">Completado</span><br><br>' +
        'Total: <strong>' + fmt(d.total) + '</strong><br>' +
        'Monto: <strong style="color:var(--accent2)">' + fmtPeso(d.monto_total) + '</strong><br>' +
        'Sin factura: <strong style="color:var(--danger)">' + fmt(d.sin_factura) + ' (' + fmtDec(d.sin_factura/d.total*100) + '%)</strong><br>' +
        'Proveedores: <strong>' + d.n_proveedores + '</strong><br>' +
        'Locales: <strong>' + d.n_locales + '</strong>';

      var kpis = document.getElementById('pf-kpis');
      kpis.style.display = 'grid';
      kpis.innerHTML =
        '<div class="kpi" style="--kpi-color:var(--accent)"><div class="kpi-label">Total proformas</div><div class="kpi-value">' + fmt(d.total) + '</div></div>' +
        '<div class="kpi" style="--kpi-color:var(--accent2)"><div class="kpi-label">Monto total</div><div class="kpi-value" style="font-size:20px">' + fmtPeso(d.monto_total) + '</div></div>' +
        '<div class="kpi" style="--kpi-color:var(--danger)"><div class="kpi-label">Sin factura</div><div class="kpi-value">' + fmt(d.sin_factura) + '</div><div class="kpi-sub">' + fmtDec(d.sin_factura/d.total*100) + '% del total</div></div>' +
        '<div class="kpi" style="--kpi-color:var(--accent)"><div class="kpi-label">Proveedores</div><div class="kpi-value">' + d.n_proveedores + '</div></div>' +
        '<div class="kpi" style="--kpi-color:var(--accent)"><div class="kpi-label">Locales</div><div class="kpi-value">' + d.n_locales + '</div></div>' +
        '<div class="kpi" style="--kpi-color:var(--accent2)"><div class="kpi-label">Ticket promedio</div><div class="kpi-value" style="font-size:20px">' + fmtPeso(d.ticket_prom) + '</div></div>';

      document.getElementById('pf-t1').innerHTML = d.sin_factura_list.map(function(r) {
        return '<tr><td>' + r.p + '</td><td class="num" style="color:var(--danger)">' + r.sf + '</td><td class="num">' + r.t + '</td><td class="num" style="color:' + (r.pct>30?'var(--danger)':'var(--accent2)') + '">' + fmtDec(r.pct) + '%</td></tr>';
      }).join('');
      document.getElementById('pf-t2').innerHTML = d.proveedores.map(function(r) {
        return '<tr><td><strong>' + r.p + '</strong></td><td class="num">' + r.n + '</td><td class="num">' + fmtPeso(r.m) + '</td><td class="num">' + fmtPeso(r.tp) + '</td></tr>';
      }).join('');
      document.getElementById('pf-t3').innerHTML = d.locales.map(function(r) {
        return '<tr><td><strong>' + r.l + '</strong></td><td class="num">' + r.n + '</td><td class="num">' + fmtPeso(r.m) + '</td><td class="num">' + fmtPeso(r.tp) + '</td></tr>';
      }).join('');
      document.getElementById('pf-t4').innerHTML = d.ingenieros.map(function(r) {
        return '<tr><td><strong>' + r.r + '</strong></td><td class="num">' + r.n + '</td><td class="num">' + fmtPeso(r.m) + '</td><td class="num">' + fmtDec(r.pct) + '%</td></tr>';
      }).join('');
      document.getElementById('pf-t5').innerHTML = d.trabajos.map(function(r) {
        return '<tr><td>' + r.t + '</td><td class="num">' + r.n + '</td><td class="num">' + fmtPeso(r.m) + '</td><td class="num">' + fmtPeso(r.tp) + '</td></tr>';
      }).join('');

      document.getElementById('pf-tablas').style.display = 'block';
    }
  } catch(e) {
    res.innerHTML = '<span style="color:var(--danger)">Error: ' + e.message + '</span>';
  }
  btn.textContent = 'Analizar proformas';
  btn.disabled = false;
}

// INIT
loadAll();

</script>

<!-- IA FLOTANTE -->
<style>
  .ia-btn {
    position: fixed; bottom: 28px; right: 28px;
    width: 52px; height: 52px;
    background: linear-gradient(135deg, #4f8ef7, #4ff7b0);
    border: none; border-radius: 50%; cursor: pointer;
    display: flex; align-items: center; justify-content: center;
    font-size: 22px;
    box-shadow: 0 4px 20px rgba(79,142,247,0.4);
    z-index: 1000; transition: transform 0.2s, box-shadow 0.2s;
  }
  .ia-btn:hover { transform: scale(1.08); box-shadow: 0 6px 28px rgba(79,142,247,0.6); }
  .ia-panel {
    position: fixed; bottom: 92px; right: 28px;
    width: 420px; max-height: 600px;
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 12px; box-shadow: 0 8px 40px rgba(0,0,0,0.5);
    z-index: 999; display: none; flex-direction: column; overflow: hidden;
  }
  .ia-panel.open { display: flex; }
  .ia-header {
    padding: 14px 16px; border-bottom: 1px solid var(--border);
    display: flex; align-items: center; justify-content: space-between;
    background: var(--surface2);
  }
  .ia-header-left { display: flex; align-items: center; gap: 10px; }
  .ia-header-icon {
    width: 32px; height: 32px;
    background: linear-gradient(135deg, #4f8ef7, #4ff7b0);
    border-radius: 8px; display: flex; align-items: center; justify-content: center; font-size: 16px;
  }
  .ia-header-title { font-family: var(--mono); font-size: 12px; font-weight: 600; color: var(--text); }
  .ia-header-sub { font-size: 10px; color: var(--text2); margin-top: 1px; }
  .ia-close { background: none; border: none; color: var(--text2); cursor: pointer; font-size: 18px; padding: 4px; line-height: 1; }
  .ia-close:hover { color: var(--text); }
  .ia-messages {
    flex: 1; overflow-y: auto; padding: 16px;
    display: flex; flex-direction: column; gap: 12px;
    min-height: 200px; max-height: 400px;
  }
  .ia-msg { max-width: 90%; padding: 10px 14px; border-radius: 8px; font-size: 13px; line-height: 1.5; white-space: pre-wrap; }
  .ia-msg.user { align-self: flex-end; background: var(--accent); color: #fff; border-bottom-right-radius: 2px; }
  .ia-msg.ia { align-self: flex-start; background: var(--surface2); color: var(--text); border: 1px solid var(--border); border-bottom-left-radius: 2px; }
  .ia-msg.ia.thinking { color: var(--text2); font-style: italic; font-size: 12px; }
  .ia-suggestions { padding: 8px 16px; display: flex; flex-wrap: wrap; gap: 6px; border-top: 1px solid var(--border); }
  .ia-chip {
    background: var(--surface2); border: 1px solid var(--border); color: var(--text2);
    padding: 4px 10px; border-radius: 20px; font-size: 11px; cursor: pointer;
    transition: all 0.15s; font-family: var(--sans);
  }
  .ia-chip:hover { border-color: var(--accent); color: var(--accent); }
  .ia-input-row { padding: 12px 16px; border-top: 1px solid var(--border); display: flex; gap: 8px; align-items: center; }
  .ia-input {
    flex: 1; background: var(--surface2); border: 1px solid var(--border);
    color: var(--text); padding: 8px 12px; border-radius: 6px;
    font-size: 13px; font-family: var(--sans); outline: none; height: 36px;
  }
  .ia-input:focus { border-color: var(--accent); }
  .ia-send {
    background: var(--accent); border: none; color: #fff;
    width: 36px; height: 36px; border-radius: 6px; cursor: pointer;
    font-size: 16px; display: flex; align-items: center; justify-content: center;
    transition: opacity 0.15s; flex-shrink: 0;
  }
  .ia-send:hover { opacity: 0.85; }
  .ia-send:disabled { opacity: 0.4; cursor: not-allowed; }
</style>

<button class="ia-btn" onclick="toggleIA()" title="Analista IA" style="padding:0;overflow:hidden;background:none;border:2px solid rgba(79,142,247,0.5)"><img src="data:image/webp;base64,UklGRr4tAABXRUJQVlA4ILItAADQwgCdASpAAVABPpFAm0mlo6KmqXWK0NASCWVuzmSokkLxPnjfDR8zyVbZvqc/sisy+aH8rwr9AP2zRmyj9uWpl4b/0/X12o/sfiO5Jdo0AT6+efv9/50fwf+n9gTv1fHN+/f872Ef6d/jv/P7Sn/F5jf17/g+wx+vPXM/dL2iv25M007H1q4voOeK6H5CoxQFTUfMSXRr3Z1h94n7QOY27POpIUphsNglOLsvYKVgp8i+ka+DAUfVLCb35/M+ZmANVbmlW2XVW4v9piT8jZKOgRMRfTbUwo/SxwGkVc7Rgb/p8k1Uan1UGja58uh+Vn/+ZTwq0A8IPmz7n0wFKLGY2W1CV3+ZYfX3Py8KgJyNHwX9lyxjjjmn/6Ybakr//brD9lBFUXX6+BdNIe5Z1fV/yqEf6Tg/aGFhIGcwQSVvjNhcrhzkFW/Lsx5Teea39j1ngn2v/mE0TQLQsHVz6+z8ZAlUqah6blga/dsjHxMrtstMzJJufUn8Xly/f11y7pkIzExVcWvl0wLB0fP0KxdHtorc7TF3VZBDl5Fd3nMGMUmxKhRQ3EAxz7BXfV47654Ge8902eKW4Ww7bB3yNUYfrGH6u8oaQGlj/oQZ4zcDcBrCrWmWXbQ+MqKuHX9T0dkEe73JXL9/R49lq1bd6BMauTYiDNV5Q0eOoG+IQp6pa7cpKIcZf+C3mrlf/8fKDD2omc8i1kzJrqZp4xb+AzKBWutjwg+0ULdgWgVrdsORuqUQwkU3SSm6A9GITcTYfsau4aT/zjFsNBGZ8NtIAdUJTrUUpXAbYX6uofnob0oClf5R2Dq5SA5kyWAkAocCHtvCDTMCRRE1g5W0QikHkU+qPpSjgTXTVgER4ZLV+ctHSjK6Ky5yP0yfPGcKvVCYlnXevk0D6KgpXtJsqo0M/Zx40AQQXPLd7eF/qfoWIkk1hDykGuWvbbggcYafLvbxjhQfGMBB2ooPU30i01b6+4nNAzo4+uaOn7ffQ05dfw2WuB1tTaCzvArZgjVfVRt7Zb+MuCfjkszXlG+QSPDHOFi1TzF47KF9gq94CDtyORzP+hpsp1ZNSAMzdDbA3eOBm6cl/hGMh/FTI4oetBy7Igzu7OAryCIwI879ozM2NyY8uUas9nkjsWW7Hm92tSa3Cuf6jPnwI6vcvbXUtcKadF3QQ2mBWH8X2bnjSF2j6bt10RfbHolMb1sFJDKnoXpXGpt8KueGQm13JTZ38TbHUQJgYeAy7roo+WV6+0E/gL9+tp93ijhVSgRK3L1R0Y1DEbndPAb9zuyeWKK0KkzsH85zKk4xOICuRIYoICVjpR1q5rO/Sm277QHzvyv6fPlHbWosp2n4gnL8yM8xWpEXyNBkxoMJ9DVW7m09mrbJx/9XKJ14vL5Hzz4YSLJqr7X5Vn584OYtFIvZ9vukyxWeyKZd7phiU+JLxc0km/uYXkLifNQevoohJB9b+5KAW9q6ztWiDFIfV9DqNp/6e0Y2/8IBvYrJS7XNnsoweA3ly/J923hqdhYCs+gVuo3QnVBBNdEesO7LQGNEA9/C84zz1ploWKjeE1sgHyjWTFyruRIYeFmRu3OWt7ZpZ1nPgaX7/AjnfsxR5YOgo76CnO2FSM9hcnmR6mt4P6PrvFOxIa9DKargbbQ4p//TeuMwfKrTeNlsTbhpQIcy9fr7hKuALQrRCjqmCTcbzfoE5pqc8oT5iaofJQNr3IFyp1Vkb5Wi4nl46BeLf7beQS8GFSsWJKyP+mza8QfThzx135TN9LXi6KDaP7DkhBWztMFBu7gUnbcWWyW5/iHaa+tjCiKy+somPFpQ0+98rSaf6NhY+LemDfWPLn5n+sL3pwUsTW0MZeUU51ri6J0f11VPGOdR5Yp/S9BAxJpcbSEg3gzM+4YJbXTjEk0BQ4d79bueHFdfNn2vr7zPi+Jm/QCqmVv5o52rvL0x2b8kCXxhGqGx65TG+yoCxcOkAYWJ9A5XQ+s0JnBALRdHnzQUIi0mjdE5FZZ+Ygp++x1Fhqw7RXvELGvvKk/X/hw7Vfs6nGT8iuNpfJwrfQZY/2Uhp1hME/Zd5nT8tqs62SQm6AbNtVlpcbQAAP77os3ZhI+fkev7XB3T1s9zmgobNZgtwKFS3i9skAAAEYFxDHixRVMci4DOjIDZkaA7OrqewoPj3bIodWbfjG/kFoc2QF9ojlDOSfpag3xjy+It7U5saKp8QUUsFNt4KSNZe2ykpD1AwffHrAQPjAgLGDhgBsyCd7UBsyJkEML3g5Kjuq+VounJNLEGUJ/2jWSlk0148NDDMmZ3H3nr4HXpdWspSrUdafNtN0ZeAcel51qug96IykMr1LLhdSJytCWQ/XRMSAAAEXZqlLE1OZT9XnJCBGta+blr8dt4mdHYGbdIjw5LJ8QNAfmFfdf9191dKL9+SdDlRPA/RcAotTA+bwAAKiZjVsz/+w1rvkaWIZbm9tERBwz5Sjh8L7Y6+CE/KTv0jaP00o/EAJWGWztKuDhPiWr9EAEbM3YT7WEj1x3dQApZZAxCrX0QUW3c67xAvSHy3IbVOoT8lWr0HbC1XGv6uFIpSk7kIzTEbX9TwP0TSRGq0PwpoW1MloHl0NIKhL83TSUFok63MHQA42pthDlc3V91QesLlDTLIbxqcVx7Xk4d4cCSkdV7wtsI6bsC4zn2g7K4Pp1JI7OUStbFg5g4B+iL11k5909BrH5scXpj3VNouiWYrIdocx7xXSPC4L0ZBdx7zWF5rQ85Fk+MHJpTmgwk80CYWFzPE1bRVoY/hlJI04I3BUgP1XE9ZSF6YySPY9HZPSxvnb1y7hCfTkIkUvyywkxwxx8Rte8T3iBjvdiPPdIa1+cAtc6kUT8AV6EAvY3/nrtounT9YwyojNcj+xJy6J4qKFBImWjGYNcZn65ebUlHc+8gwF95LXYId8+WlQzHwp0I9ShFzoripmHckgjFqR2b2PH8wmj5TD69AtwfWBmedaUoOOhFwtdRyjCFMKglQ0Z3FuYHdMkL5uoM17mxmw8/0PecvDweV9adazkLUnZdYjktKzoRBmwx0wiwIx9uL4YYVSK2vI+UGJwZpPR+rt7x9PVTnFSnY3M4Zd27D/j4ArLtFhpMKRIj9zctHg1cJyHyz2Mrx1Q2tzF1dre0gQDKy6PTuHm3kZf29HXfEzi20CKBGCntPSd1wu7RyBfoKJxBYQNtDqNxef0/olR1j9KVvPCyTscuXbBTH12FZM1Ui3jcwOcw6lCX7e1CuIaMl03FnBYaQcDz5ZMjSl442+QsuqznFoNsikEGiVk3kK0JQLKGVgF6h6tlH22F+m4mxxJvJ0hvU8x7bQE0SgQCJab534uSAbh3N4+U6DB7NBQ35OhyeoAIifCxS8D9hpoRNgn1K+GihU0SXiL68cpTzTQQmGfAdaAG7/NwkljUJyqfuJGCe3Jt1wzHYXKOmW44wYrDpIWNcKfFoFbnpDjj1DBa/r+Ne3t3syMl7LNWhBaJVeB7dM1m09+vR2LaXXJL2rc9B4FIsretomU2dVeLUZR8aunPiY9kP/KgcTsBk/cNyGUShcmBicjcoxuQb2+F3EFSdxHx/1WuCmeV/EEs9Bxpt7hPCym7rRddcMChxhEHxuIXsXpo3DBu5cFEDgqttMV8s3klUkbcSq4DhevDA3d9YSFH2cq0GkZYrcXuWfjM5bScVbn5t33p8D4KxWwIT6fdrvsjcZoRxM74JVaBBGoWuSqKMuJcyzPbALK4dUUg/fgKTPaLpNCOpMba4gqDFXiK4DGX2Rj23eVK+JTeNz0Zitz6iE0iFqJfCLKxRZedco4c08pX/VEve2Y+W7HGH5phKmvbexx9XSnPUWcqAF4/yY6LcFDZjZQmoGzWqpzJJbdUKlWFlbRUDbSqfJBryeOe1kb+1FlsAUbsM3gBzCpKVcNTnzw+6yOY4RgRJOR4LaaOD9vIURvN6hN7DcRci1IJPtxTJsHSW0Xp2sgd0GABkZ5yd1fFgKe4hGONiZUdEutqJzOHkFs5/l6eue8ALMUhgpm94IZ/DK9mySwmM6/PKoCgnk3qthhqva7BVGHRr8f+dgjP3Muv8KQ9Kj/SpLiuJrF1MtP8dWPOxXyyKtP5bxdQZtaUaw+ssuWhdPXVH7hrEz1AVzjGJJpoVE5Ba65A6RjKR0RuxivQrIVkJBP9KkqNjZ7eMVM6IxdwyyNRsBf/XzvIyVWK215CkCZHjeAifl1ravcPSUtkqBQU4KnMU3LexqZ+BGx8lMqjCa9wE0ohF1G9Rzf5L5QpVP1E3LewOI3GYHch2nM+vgqU6m7+e9XPpM2btZf4QgKevnyX2pH2G/HhoXdDAH7nk5lsSZN8ISKTSNvCtqO0Lxluz0CzmfkWGfSwwk2/c9BbQ8oSVsd1LlzVcJBgkYQ5JdSeZbTVMKCCawDTw22lSrttA6KNLRV8z0D9V31x0UatTc3cmxFC3liAL/waIQvdE1B++jE+JYGvWHwQ8rbw+Z29AD/9NUYgq7tdajWvqn6EE0CewUcOoLril5kJ/hXCMAmy1DkvUVV6NqjNVc8w80ReUHbNu//Dbn4XyYHvGBfkalWw3lOdPRcnAH1J24NJvjQZaQkZ+j93Bqi+dbSrXrU3I8nQRiK77X0JI6zEGKBOt7Pj1G2prk8TjEDjib1o3/8NPLx45AdVlrzp5vbA7ZL1yCxTnaMmXoEBljKuj8ThmBDJKnixFWi+vAnizcrYQ7VMaU71FxD9VF6pXpuQnIuKy37FfqItiKg1ynvvQzXSqsugFyugOuqi9kwU9Z9C75rfMfnqPuTmmtMunfKtxqYpdXoLYWrjT6bmhBSY2I4Tf3sAqZ3lewJTnJsvhoKuruCwmAsZmlhoKna/HCRCVIhGW1BFLkG3ZUQCtUiHhiButjV75XzXPkv7N9r9ARyLGJn1COXvpxf/Ao35EmzX5FhZoWnHGQ0GUTEecAUFcn2HsG2ROW3+TqcXj1/F94Z4VjsRE79H1POG3dtcI6Ny/9OwYTXh/D41uuIdXOUaCUAZTQdmdMpKdktoWnkud42GhjRGFy6DIm9US5Mr7LbPzMHC8/Mapn8u0v2cshI4vCraW/kSCYQgJsbDu4xJK/g97LqjwA5MuBjCA9DijN8DeJw4S16d32oQJFFMXev6rz2rk5XBVW3Ll+kuelp8I3+veYZGScbXsGYrTrl8OfUVZ4FxPt42pxsgulRpjfU6qqh0XkWVayiG3JjM9vFsVVNYO45872VumoTWDse1orFQppOHwHtiLGCA7SSICPpK5600XIJV2WcVBUud6TO/J8eZZz+wzBdvVThhHd36A8m9qMZoN6vro9oqjJJJMjssJSYKKc9rdjsuf8hWlDMWv5FO1QrRug+5U59t5eVVJuYnnyULGgR7mLmvDxwp9SZhTyWyOuYJO2483/nKcrWhD5m9op/0U6URMpplmd+W0tPDbYbX+cXy48GiIQo43mzboq/XwIH4a6jfB0sSxLtwRea0Ojwk6j9BkADxdaS8l+FnzMwLlM5t18spK53QyPxKwsdyLXqZRmt0QirBAt/P+WHtptbG/+niaj6WcNJQKVJ9oEbZEhRKzbAEYuMePtkzpVpw00DjaBHK/oVhBfciA09dtErEC7Uy3+99S8QYKNsUJ5UTkgpip2W7VxT4H63IWDwK5EH+5dHFxbZVvPf7iAVXRR7dFnaZdW8A3evEakYTGobOKL4MZEjZAZ9SEE+h51KET0Ev7shcRevTB+BIQK3g32TMwjSmR8Juxn/W1oh24v9zRNVYyFJuKldgVmc5bRudVPvLmveCXkYYEno9p7gckfOk5sKnkL/V3jpnoE/NXq9m6vHoGvFCTXKt3ONk2Gp+4Au3tVHgYNyzCUPtnbc68Ig3jdYkJQr7a1axiZrhdVyF0YB2kBvckIkV9/jIWSgDaBlJ0bGJXoKQvWrf12EE1anfLLSRK2wg3HnXl5aalAhQRMnO+kpUfj4J7XOE9Ly0XBCg4AjX9A0MCp/c2rRtfNcoj/a0ggBGF4bqNixnGhWlW9JFKczbnmO76cILh7bk8/4YbBsvoS6dcaphAo9Qg/CB+qtScfdIQ7W3kQl3OOH+s2OlBHlZUL22y50Wm6EFTT4cpXzg0IEyfAXdcielIKqhx7Z2pSh2yUgXgLwkbEENeg+y8R38KiGHE9l9gnqMGOe+Q44f2z4JoAL6Mja1RbOtQxiTTbs7nj+f16fW2VIpDQy1KjsRqHpM6ySfNqhdYhxxuIC/oe8URLhhxbgLBB/MqAlVnJF9HlRIa8xab1Ox4XWD0NzjQ+ZN7Eg7Ex4tihsIhNrKlN+nwvc4Omv2lTEmY/WAJvYyVIwC6A16bmTfCGm7HRhG93SQUQqiqLrZENHkMeEtffWPE8eGjXy/I9Vd7GlfzkP4Zct4dYdxOa8e/mgn9sv/xBRaZvL7lcGaY+H6SiQDkeKWTfs+aOc5TBtjjVy/mN1jn4FPi0a4ts1d/jLbAYQm62RhhFF1teoX6LsC9c5YJHJrfZymzTu9qAUbHQkiriYR4RulXd5HtX82hXDv8Ae56bDTZOrT6P2YNz75yIT81lu5EBHMexUW+fG1fTHG+H4x1W5XZu0krdpBe3QmyJwccvDdJpapQxmh/VayGEAva+6Q3a+60N7B4ovZUfWVTPPy3K9ONMHUVeV8BqzogI9t4vDdyR0cYkHsN2plNdPVLMyPuLAip6aaeEGA507w98at7kaAEHfnUoFHbdRkmux3N7O3WCMu6K+aPxIoO/QGrGYh0Xc6eli3qJ066H3ISnPLUQtQj/WW9VZUcyTYLCYwwnj8Y1fQC/aVqKTXgtz7BV3K2uCfnVq/IWNudne0O8v4gkDWXdlgsvdyZOXO0+NpOS6afrIVbvlM9KLDF5p8qwYKjRXBOv0rOflVmEbsRPFDZa02zMQ982sYQpOq50GjDDTE2ktJr7wRuyQCuxdoRZ7imF/KSqcibViWJdXbTDm0lL1xjTBMwMIkqSuRzATeR3vbYJJCkLABOVVeiXUpk4ENRIel+kaNXHK/LnHeIsNMnOT6TIZgJXq1Sk97Eb7PgcHgVrF5xUh5FaXpATZ3jrp49fLh1YtMnox8mPUlZgUoNuLxcCFsUVgKzAvWdFHayKqDAmW+/s499aU1HyojqJIbpzK1fYruYevDtjrosBI2ibH36KVemOKDIEVPiiDZrLptSFWWuu+d4ChY4kOmh0ZSOGUdHz1zi2mTB7Sdf/AHGywPqoPBHxGFbMbAq0bbnsnL1yyrCoshUS/kZNg3ymQIPdaVwC8vNU38G23+8OoKQfVeH3CwQoVsajhYK8EGYjlet4nU702NaKSKHI2nlK2rKj9wdiyu+odWgdlVdD0m50Zl43bqjBWxKsevv61pOJ1GuasjAJLm98eBdk4agqMIIKa1Oo9bQnQVCkRYGDqsucHWjed0OzHZeIrCjXKpRG5g9NkMLZg9uePwNBkr9SoeqD5zQeKvL3WB2k/i7qDAeiHVOCPlhUvyRZ2fASKtLI4ub2EFpDjpEZ0+sQiQZFhVZijm3FG6mqKBAAgTHK0yKBdT0cQfW8AuRzJ4ga5zL5AUN4Oob33IImdVn3Kc3aVJOydys/CiZkoiRb+iF+Cb44uPfM9QiA74KMUmeDBcGmdygsIEdhkdxeK4m1fPZkh0ofnzD25bJijLa1VKpE65Fej0IOoWeEz6qxyKRxhV39FSTdomaRdp/gPiKB+aSMq9BmbSMF22s8RNNE9AFXYmnLuVmBemv10Crr6CdeqzGPDT1g1ffRMsadEz2FDQoB4rOTfm9XOSAeo0/143oZJ/F1IFPSaYU6di6NIWX6/lfrd1toRrF5lYIeydA8eNXnD9ov5bX4Kr/twNHXjLk6lc5D7dutbruW2sZ4fKFPpYdjYQDKgNLi1atq32GSusv/XSwJi3et5Nh3vwuBWrGaZWcgi/r9178o17TQy8H2kj/9FNixeW2saib+co2IF50zF69dlnS0SwkGSJff+Ndi2tBay/IYesRE4XYzIvymiaJ8raJTiPfK34hzCrAlNMaJYC+cZjB+9f/RhCvgWJ/j7zsemQC24KV5kqI6Va08F4RWMN/psgwaRq1bL5B1YC7quS2BaysUWG9AWynk6FCnQbS/X6ow0NA4O+hygECESCqmX4bt1jWoZT7ZBjieuOd9HL303Jh9c5d2FYFwSI/PVE61GSIkDyOTj+TzG2STPd4iyyBSz3rs5rmxR3Oy0TSI5LHWrP0s7tIT2mcEbcNUnJI5tr/Kd2Z4JLA3MvtGuEc2r9wvMnhHAs8eef3SW0fcHZ904OCl1d2TPKunEpATQg2Hw79CDkhsUmUtQR4ylPo+wk8IdcSfg8Iw7HbWZ5utVNzpfPcTq6zPh5zb37rLdAva81yl6q0/CcKmYfMBUFeGHoXKY+jy8F20iNefBw4KT7VqSRBkV+cfRVPGWueJDWfJy/WAMgyP9PJjil6+PqYbYbOUSdTk/Z/ok5tmxPbm55cHZcQh8hkIZDuudcsDV5ioo/sztsX/vAe0tTxWBUsR3aL1nl1y0/Ask8KIsEMXXNES2eZoNXf5NIStjH8r1EpdK1zIGYzFEoADe9jipiZWBmXNUGmsrm9cpl2DSTF8xYd7VtMDLiYuPkOdCs3iApx2324zUlp5xigzy+RGrgXSnn4lFSrH5MTO3SipO7WxER4TpMDKaJKvTXZ2OoTYp72TZTBQaJ8RV5sqO932jLOOALv+pI9PXmzjKDMVVGSCorRxhu2CpOfMMP8vCd+JacnZgdGfDg5zqg0HVuPsePhnmHcx8Rl9CL5qHwHgeQsQIygDmAEVpCK1gGfsXHc1SUekpWXc6efcfaY8m5U2C291NTUNoEpFbt2kM90sa2mLrS9cAGRem/RRjWXhYB5kwFwtGrk1esZGx8htI1fC4j+SuzPlRFZ4JkT+/PGPURgKxPCLR4uwNDiH4VFuMDQtabxcuno5kGCpC3t2365lVR1sxYQpxh/++VONLEXJmfBtAkPQX/qImXbEuCnXrH+IMMfnV4n+AhgS1enD2NQn7iDEyNuCp9k4Mot66+h13vj9nyHh9fEfP115ZQbzJs9N12mW3zSBDFNuf52ZQeb/YIi5TZDP6lRVJeQX/dm8p7BxXdbrleX4HWjKWVONoYONHi9xtnuzpn+MOZml1mgWa6GAEdmYDxUij/slicAJvCtjWAmsllNBE/l9ZtOfKebYjr6Sbn08uT6oAqoqLjfWkWphk0hZvtKRvBAmHXRhqN3WKW8+1ckYbbXyB7fuCfG+sqxKx2iSqfRMVpjQHMd0Md9PacBUdl66NAo0VLUd6OPgbkSAQU9s19gjjBMzNgg8Fmum5WL7kNx1+4ASS+Lj+9Y5gFJoSVKb5+ZmsgeeSiP3yj0mRYYOYQnm88PZuxyrVtVyoxI6ISBtZuIPK2OUoBWqeXzcgwqRtchssHASEvXz/BpDVIh7DrjnPHQ81z3t3TpFW8HYtEl1L1XKKgrklqoUq9zVXuv52Qi937z1f3MeJUlyV5A+/LSevnWICMZXBRqKOrD178RN3mG8AhCjHGSlNL7miKGMe5EwKcrjnV26wzR+nye7WLTUyijrTTfQITYapv2V6P3TrMm9HpLVf4UwDL8PfY5FeJuQIvdC72+7kIHH0+YRp9h+zqJTH7ADNHx+BIgfYDY/aEt8A3ZbefTUl5sbUP3dkbccyLJxBQ9qaNOH7+JoMy373KJvnoMq3VzF8byvF773PtnvIhjTSUHM0OKGPpva+dsoCHQuYRB2tnQ6ROihHx2Iljqj45ZQeiqGmojLjfyQ4Jp3fiBXkyw7Wfjj5vgUmwaqhsH7F9vit0/fWtzWVURl7B/CS12GSsas1qZuUvIAGMLmx1tSiHr7xipxtY994eC8UEZtL5m7vrG5OM/xaHqXxo8VTfP0TNi+atkEESwMo1uAWfREaWqdY4hXEhk3cWx9zg1jPOqfnJcnlQr29Dmn64/431u55sWsmEaSi0E+m7/PbgXQ4eYeKxSlzCEFKAHfBVuslKgnlrNSAVQyvF4jTdP6ADjzNwEKvUbvvlK5InPw13minVxdr/IGEA9bXQnnKMxkqLoLYXud/TBdksNlCNvKBiIysjQ60nSgFJ8BdYfKPXMMZSlmGkd1S91LSVFYSWGNoY0YqDkdzmrK9KTrPArRdLguuGequf4PadNjuBoxlmLzLTCbO2/Os886nIEhuLFaRhygj6p2NAnXlH7UoiPNuEMrNcBXOXwtM8hcB9uUcrF1z0GJ+L7ceja9x/r+tt2LeC82E2+XJ9e+BnKAln8m1YGU2QxMIHYfblrjLuxQr1cz8JbJCrM4tJRI9ZNbmn/3CQNUf640Hwgi+i3xqyb2DhsjQ0IH5zaR5vHlDseol00M3F8dJhKGp0P3Xe9t6iiYefAPmm7Vmrx8gUbmtxqoO2EkKwZPqfeng33/t/zwbyYuCvWeHyDnUHYe3U4WAiiBPL/1oJd5Fn8lB/9EDjx8URLk09WrIf7s5Sde+1GVoT/y31Km3EXv/YrjqqKReAT+qWpSJnDtw2mcs1Hz2ktjxIGSPpEVaNxbtvTbEZA+mReq2jCToJkE1AW7xb7K8H6PGtOE9R2xVG3hMMxYW14jrwQHX/Nogah7TZxJwMQ2WdvmasjZYbj91c8s8/iv5YOdENt+PUc9MSxqliXHkP79sP27Et6/TKo0shLnsYLacruEYX8WOoupfqEr5um4YKDCFpObdWtcwH5mNgxDI0R4P0YIL3rBZGuTMIGx+K0lHMbXHidVtwojtyWWdCAI9Vua+/UeSHcLEDbr5qMQzR+Z3SUidmdNF+SkZxrj5yJH0nzqVGLjadcifmv+f1MgaPh/aqMvLejaRliP3EapEioKG8nBqGeLzggLUKv41Mt5hGkjnmir+dINoGx9xPqvi2Vbfn/5zI2KxvSL/dfJm7KpjKU43YZAdGROTsUYcPcbJVqdaOREKLiHsxGvtfdu676apgwgSHW7p/jC2eJHH1wfX632WMdG+MpJs30xuyNkm3uqW9DQUpo501Gt5rzO1QB1X5lewBcySY1AprN+LSeo4zThMrZqMbIAczmwNHBsxrwbwzLffSU+2QBrgpxdBdFH3q0s7sXoX/e+vOHld6LrP4sU1CubnDuZTTj6HJaGYj6p0Mzmo+ET9wL0fmiwz8vpepSZLOJGD36RLWMQV1JVMdJI1lrv/CWm4N6VaY0QQv0LiyKb21RLgb3ZEKvofE+ksInEQ4cmsPiXXUGwGaFKp+JhdaalSPn3Z5zmdL0xY4ngPwAw+Y92rQI+ZvWAF3OIgmKl7YfAF9PRQ3c/PywZ0Z35ulMwTAutqexx+oe5EgZ6kkLhm2LPfjhbwQG6PRCnRVEWlZ4HNnvcXmdl9VLPZYFPFB/xxZiojgnAzfI7rdzPlvTb938Lt1O/OXLc4exxTQV6iFTok5ED/paypXtCt8nL3DKfiy0bCn0qDP0Ry9+HEYwCMtPV4qw7BsZ8SMDxZPEXq/GlDpowLwWPYAyMM8b1iaaPDsLrAB7gGF9MZ22QQw4bms1yga1gCJ0VYB3BRxySnBnMcNC9C5OeXTrgpSSTlZje6oTfDXuSgnl5jNa0isoQsa66s9HSJ2q3V1W0ovaQgPRtBpIfZRoqJowg6tjH+sPEJinc3Zo0Y+4umK7Oc3PFZsfTM/S87sVOZi+DKtyuyurKDwnwFwQrrSo5Z5oV+6uXoIvFwvD19J3V/d8s0rf6Y7Sq+9znED2Tv/cAusF/cl9W6PmNU/dU5+WiX/+Vc/mxPOfIpQp5FP7QpAa4xWTwmReaGRxv77KLwTo/y8q7ffmA2fw+pAjJc83JUXxjWU6MEMgkAf5Ijfz8kDI5/gSzWdpxYyz000MNr/AMZM3EQ9eTtG6CyLu5EfHvT7ErJIuyVx4KsI5inN0i8UbYFS3jiuG0JcQoDPQBMvx4D8eSyI4YzHqHR8I50l5qUslOPw6M0+/pWJm27vgXVUcVG4GxPcDBvA/9Ls2xeyMZ20ZJHdeRJ9l3b1QQvdmjq8Y9NE9vJ2BVXO3Afr3zMEu5/fkxoTsgxjkbdGvu+cRIMyqvMiJnMphE1Ct0AO3bZw9VnGEXjasP/yCDphO6QN85WFOrnykNqRTRNU2uqejv/wAK2oCI/g7SDh6atfPlZ/4yM5p8an6qu6bfgTaYfEB7/mKJZJfNEQbY0XbxWbl+7le/e4k+v3BnqQOr3YvBzEHVHcJ6No/P72WGqhEYs5GKuzTLmBAJPp+qaR9I35/yInBDFPgsBtnX1uHMfpDvI0YIjy4rq3+1felnEHNTDDe5Mp3K55zhFeE6SAFq8K7pz+3CxnTR/sHnCYMnGpnTVcoO8SJfRBgT7x1YmI/zpsNZs5vqSWL6RD+9qt5rX+yS0+rTZmUekKra9Uih/tNI5n2wngzQYPtexC7SaKSPwvHb+j9uAvXujS23pfY9Y0/K9LHA+Sow1Dv6qU6kfGA59oEhgZd4JsutrTZBkfygfuc6axgjzkKPeiAIBnHs95ekajPOupQyza9ruihaIdb5251a0BlKqde0iorJTM5YG4YUcb6ZoKOqoNXudvMiNqVTe1LjBmLasr/4UAuXy/LSj8NMA4w+z5hHb8Ra1MYqZbbb2To9ekY/6ePkqtf/kpqLHeVrtdaecUZqEsaSW86IQYC+eVcsvuTwPwY3ntnsngCLDkChOToyuFI1bGkASPnfYHpEPwqoCwXTs4GLFVaJXvjzEY+9L7/SCV2DuEfnkjTUr3/g4QnurCRN0D5TC692Lcxa/ADTD/O6LemndEb9dhnpId9EKZItnw7hDTZ+BOyTjuwY5r7DKLgLAy+pI3C3fB/7RgsQjit0lNgYgfI0d78mxQKPTTZGolWR4zKQFk2b5f+ZikTw5aLcwz0plk2F0KRNTrQmSkfp4AvIJYX5Tpd1edCONfSISvWhIpCTW5tdVDh+8wzRISegFb+8fk9QooyxKQzynqpzL/4dR030d2wnabjHXBzVkBRogthTFlC8LRo9yFzlM08YAKpE6F+K0Ad18Rwi/CwYX3AUjOl9Z8VtQ7qGjn3Hf5u5cBU5y/odEzfOPFzO4uhUWQ8apk2T/rWQQWZGZYNzZPjLefRktJFT0oOCDk/nSwmlkHiLGoye2Z5ZPtMiausFiyBw+jRLv8A6lNiIk9sCLG3KDCsrFAhgLQg+MhO0M2+FBvT6G0I963Jy2FP5ALQJYPlb7+jYXR2TdYVSsbPxs1MGQ9ikYwYUgV7XAKomYuOiqKO9MdPjPEyruYbGZVbQ3p800J+uskT4MbjP3wex7GdKyKdx5KvF1EXWmRoh0bpx0YPKIn2Rw6CiPL2jdHx3EYPRZKtprMRbkgtO86DXVPq4/dVOKT5oHciXBN8fCYYI637sto1NowCtI+DpwGPMKR8Dc6QTzJxb9gu0sxfR6Xv7CXBSJpgGV+0c2lyVToiKa+OqwUDDK0KAF/d4vJtRB+rbinFU82/LxKUwlEgoUgahDKY3gZgvG5yQl858saO6Yz1hQRJJ9rQQmOQ6WvoFGvmOD2a9wKO7iDT28i682RhRvq5wsV7UBMjYkAtv+2Hl9B3wU7gvScAVVKHCBn3kARe39+zRfrIzNU5pnI0veaZ2YxOZ4DBgd+CFs4UvTy+V8p4WEcSnuvs2yuNI4IDj8cjQTVMDz3X94C8YxGvWLUxEan3DURBcVSC6BvcnGHrLjNfTAvgQpQby3ohEr+Fo8CrD9LSFZvTILLzYeVOsFsfr+kMUp3SUJ+uMdkH6YBTyompcflFiIsYc47MY5MQ35EagpCwgPn95TB4S4bYj7V9izjQd+DKfNcqa9sBD4bkowY/qs9jtpG2eR2yc+lGeiCeoCkri19McSg43h5oXhCAuGVvmR7wrlxg2TxBmSw4VWNUrd5gbbQ96AhEU8wxnagiE1hzdzIN9b13AEaCsftg97hGjG12Szs5PYRCP4cUCL/lV61hkkErOyxwoHe+SALiigSF8bt3hZkoAXWeMlw88t0NwYq9UG/JEGLPvFREqFM6bNDEDbZGN4JrOduCrvksSqE1F6q4Z7/pGnF2TaHsjKLhJDAkOnA8e6SQzWr3fjjMoG8m+rkS8iYMdsh9MOCdOAobK0hNXbiYNmyioNV+tHorBwiZQ3eQcXQfq55aKwdxUiPJecSC9gual0M03vpVN8R5eDsIRbRYC4LaaSLRNjetrv8EyfVC/MzuYD6uA2JxaDVNBrfPLffxXoRwRs0uHQqABPqP5ndaG3kPcLOXEF9/x/U6JwSdI9aMS6g5onP9MUjpqTrJcihSeILvOlYvW3ONS1Xt8Yvqk1BXKyI008RUZqdpnw7ujqZWEffjjaIVVGjfdv26JpSraYIocwFLewXwAeok3aqfG7LxXB0Knlze7+uTIPE4+OGnXne4SaOLz+DUSSj5e/8Hfhob2N9ETk1H9K8fiLgY7kCt24EAsL3E+E2ERDZg6RWVwahB7jJyD6v56u2bKiOp7CHDxLiPkvlUwY3VtHh7AviGeP3BlIoLm2I+y1yjO3DTwV2/0PZu5VI7ySsNqtFw9yH65L6o4Tohr136MSLAh/3QkK0wWMv1vYcCDnS8gp6OSnPMEG7o7mlWj0ig1Ml7j+xoiFrQRnXdInV46x/iJWvo2P78omd1hPiRHm/fR5G00isji0fLCy8qcQ1SPb/t9Cc4eerMF2qsaKabL7sbpo6fRvicG0WJGFcO7YOzjnzk2dr93iAKEIcRyzmbYiYgLGAl5zELEUfR44Z6kvsu3+HCsXml23nigJzLAYlTeqaDSNQEg/ORy4JT5JKjmYreHUzmtK8mT26eZOAn8R5M4c3J1eBk/cSq0tQ00bb26eXjfbVYhkm2ilTh4lq/pEFZNi0TY1wNHByEEAW3lph2nmoeB5fiZHkLtdhwMtngHnms7Yb/8xu2RxNZ/YEq+RHbYkx6c1pQ9ca45wl/Hffmy56m2Nmq6KKoacQTB4p2Q4f2OppRkNigf9poI5554+R30AXGlp6xgOQfT9bj5ZyJdJzAylbhjcAedXeCGWfXSilqnpdAheZtnperlVveIVz0aX7h2Rx9HR20AlOQFfo3asNnDLfr8QrduBAkxwK0qQ4tHr4Zml54QrkLr8wvZR6AzRAEyPQ4sT/SgBovFy5i+wUIsrccPXfgbQx9a0uSsR7KP4Gwvs3s7plfSdmC/1t3/iY9Yr5eTF7e12eXrjugR2X0CykUSwGS7Fp0GZNQX72i9WxQ+HEdjDCurzzKFERzN8e3QE114idkazyJ2bxaHGweghv4I4lUaHKoLbsZu9OOOdB++YmH0blkK0KqYJGT/e3+scNxokt6iUWUwaHi1NWcjFk8dD3Y8FJphT0I9sigHTUERnhVTNV8GYA3t7PvRZf0y0heem+VrHx7UOwmuFpWwEus+8tFoQ+JRQdXhztY0IEmJG7oSdq/Wej888/gqDoFeBrnwkYy6NeDRdj2XwwdONo+HKl7d5hhOLKNESvBTMqqwexsCUMtWwpbbn/XbwVXEMpFW8htda0oFE/ZsNL0gL7fg+TzyA3W6tHTjxj0mDC4HYoOtLOjIBElfQwje8J00oKt8d8SLBXwecUAAAA" style="width:52px;height:52px;border-radius:50%;object-fit:cover;object-position:center top;filter:saturate(0.9)"></button>

<div class="ia-panel" id="ia-panel">
  <div class="ia-header">
    <div class="ia-header-left">
      <div class="ia-header-icon" style="padding:0;overflow:hidden"><img src="data:image/webp;base64,UklGRr4tAABXRUJQVlA4ILItAADQwgCdASpAAVABPpFAm0mlo6KmqXWK0NASCWVuzmSokkLxPnjfDR8zyVbZvqc/sisy+aH8rwr9AP2zRmyj9uWpl4b/0/X12o/sfiO5Jdo0AT6+efv9/50fwf+n9gTv1fHN+/f872Ef6d/jv/P7Sn/F5jf17/g+wx+vPXM/dL2iv25M007H1q4voOeK6H5CoxQFTUfMSXRr3Z1h94n7QOY27POpIUphsNglOLsvYKVgp8i+ka+DAUfVLCb35/M+ZmANVbmlW2XVW4v9piT8jZKOgRMRfTbUwo/SxwGkVc7Rgb/p8k1Uan1UGja58uh+Vn/+ZTwq0A8IPmz7n0wFKLGY2W1CV3+ZYfX3Py8KgJyNHwX9lyxjjjmn/6Ybakr//brD9lBFUXX6+BdNIe5Z1fV/yqEf6Tg/aGFhIGcwQSVvjNhcrhzkFW/Lsx5Teea39j1ngn2v/mE0TQLQsHVz6+z8ZAlUqah6blga/dsjHxMrtstMzJJufUn8Xly/f11y7pkIzExVcWvl0wLB0fP0KxdHtorc7TF3VZBDl5Fd3nMGMUmxKhRQ3EAxz7BXfV47654Ge8902eKW4Ww7bB3yNUYfrGH6u8oaQGlj/oQZ4zcDcBrCrWmWXbQ+MqKuHX9T0dkEe73JXL9/R49lq1bd6BMauTYiDNV5Q0eOoG+IQp6pa7cpKIcZf+C3mrlf/8fKDD2omc8i1kzJrqZp4xb+AzKBWutjwg+0ULdgWgVrdsORuqUQwkU3SSm6A9GITcTYfsau4aT/zjFsNBGZ8NtIAdUJTrUUpXAbYX6uofnob0oClf5R2Dq5SA5kyWAkAocCHtvCDTMCRRE1g5W0QikHkU+qPpSjgTXTVgER4ZLV+ctHSjK6Ky5yP0yfPGcKvVCYlnXevk0D6KgpXtJsqo0M/Zx40AQQXPLd7eF/qfoWIkk1hDykGuWvbbggcYafLvbxjhQfGMBB2ooPU30i01b6+4nNAzo4+uaOn7ffQ05dfw2WuB1tTaCzvArZgjVfVRt7Zb+MuCfjkszXlG+QSPDHOFi1TzF47KF9gq94CDtyORzP+hpsp1ZNSAMzdDbA3eOBm6cl/hGMh/FTI4oetBy7Igzu7OAryCIwI879ozM2NyY8uUas9nkjsWW7Hm92tSa3Cuf6jPnwI6vcvbXUtcKadF3QQ2mBWH8X2bnjSF2j6bt10RfbHolMb1sFJDKnoXpXGpt8KueGQm13JTZ38TbHUQJgYeAy7roo+WV6+0E/gL9+tp93ijhVSgRK3L1R0Y1DEbndPAb9zuyeWKK0KkzsH85zKk4xOICuRIYoICVjpR1q5rO/Sm277QHzvyv6fPlHbWosp2n4gnL8yM8xWpEXyNBkxoMJ9DVW7m09mrbJx/9XKJ14vL5Hzz4YSLJqr7X5Vn584OYtFIvZ9vukyxWeyKZd7phiU+JLxc0km/uYXkLifNQevoohJB9b+5KAW9q6ztWiDFIfV9DqNp/6e0Y2/8IBvYrJS7XNnsoweA3ly/J923hqdhYCs+gVuo3QnVBBNdEesO7LQGNEA9/C84zz1ploWKjeE1sgHyjWTFyruRIYeFmRu3OWt7ZpZ1nPgaX7/AjnfsxR5YOgo76CnO2FSM9hcnmR6mt4P6PrvFOxIa9DKargbbQ4p//TeuMwfKrTeNlsTbhpQIcy9fr7hKuALQrRCjqmCTcbzfoE5pqc8oT5iaofJQNr3IFyp1Vkb5Wi4nl46BeLf7beQS8GFSsWJKyP+mza8QfThzx135TN9LXi6KDaP7DkhBWztMFBu7gUnbcWWyW5/iHaa+tjCiKy+somPFpQ0+98rSaf6NhY+LemDfWPLn5n+sL3pwUsTW0MZeUU51ri6J0f11VPGOdR5Yp/S9BAxJpcbSEg3gzM+4YJbXTjEk0BQ4d79bueHFdfNn2vr7zPi+Jm/QCqmVv5o52rvL0x2b8kCXxhGqGx65TG+yoCxcOkAYWJ9A5XQ+s0JnBALRdHnzQUIi0mjdE5FZZ+Ygp++x1Fhqw7RXvELGvvKk/X/hw7Vfs6nGT8iuNpfJwrfQZY/2Uhp1hME/Zd5nT8tqs62SQm6AbNtVlpcbQAAP77os3ZhI+fkev7XB3T1s9zmgobNZgtwKFS3i9skAAAEYFxDHixRVMci4DOjIDZkaA7OrqewoPj3bIodWbfjG/kFoc2QF9ojlDOSfpag3xjy+It7U5saKp8QUUsFNt4KSNZe2ykpD1AwffHrAQPjAgLGDhgBsyCd7UBsyJkEML3g5Kjuq+VounJNLEGUJ/2jWSlk0148NDDMmZ3H3nr4HXpdWspSrUdafNtN0ZeAcel51qug96IykMr1LLhdSJytCWQ/XRMSAAAEXZqlLE1OZT9XnJCBGta+blr8dt4mdHYGbdIjw5LJ8QNAfmFfdf9191dKL9+SdDlRPA/RcAotTA+bwAAKiZjVsz/+w1rvkaWIZbm9tERBwz5Sjh8L7Y6+CE/KTv0jaP00o/EAJWGWztKuDhPiWr9EAEbM3YT7WEj1x3dQApZZAxCrX0QUW3c67xAvSHy3IbVOoT8lWr0HbC1XGv6uFIpSk7kIzTEbX9TwP0TSRGq0PwpoW1MloHl0NIKhL83TSUFok63MHQA42pthDlc3V91QesLlDTLIbxqcVx7Xk4d4cCSkdV7wtsI6bsC4zn2g7K4Pp1JI7OUStbFg5g4B+iL11k5909BrH5scXpj3VNouiWYrIdocx7xXSPC4L0ZBdx7zWF5rQ85Fk+MHJpTmgwk80CYWFzPE1bRVoY/hlJI04I3BUgP1XE9ZSF6YySPY9HZPSxvnb1y7hCfTkIkUvyywkxwxx8Rte8T3iBjvdiPPdIa1+cAtc6kUT8AV6EAvY3/nrtounT9YwyojNcj+xJy6J4qKFBImWjGYNcZn65ebUlHc+8gwF95LXYId8+WlQzHwp0I9ShFzoripmHckgjFqR2b2PH8wmj5TD69AtwfWBmedaUoOOhFwtdRyjCFMKglQ0Z3FuYHdMkL5uoM17mxmw8/0PecvDweV9adazkLUnZdYjktKzoRBmwx0wiwIx9uL4YYVSK2vI+UGJwZpPR+rt7x9PVTnFSnY3M4Zd27D/j4ArLtFhpMKRIj9zctHg1cJyHyz2Mrx1Q2tzF1dre0gQDKy6PTuHm3kZf29HXfEzi20CKBGCntPSd1wu7RyBfoKJxBYQNtDqNxef0/olR1j9KVvPCyTscuXbBTH12FZM1Ui3jcwOcw6lCX7e1CuIaMl03FnBYaQcDz5ZMjSl442+QsuqznFoNsikEGiVk3kK0JQLKGVgF6h6tlH22F+m4mxxJvJ0hvU8x7bQE0SgQCJab534uSAbh3N4+U6DB7NBQ35OhyeoAIifCxS8D9hpoRNgn1K+GihU0SXiL68cpTzTQQmGfAdaAG7/NwkljUJyqfuJGCe3Jt1wzHYXKOmW44wYrDpIWNcKfFoFbnpDjj1DBa/r+Ne3t3syMl7LNWhBaJVeB7dM1m09+vR2LaXXJL2rc9B4FIsretomU2dVeLUZR8aunPiY9kP/KgcTsBk/cNyGUShcmBicjcoxuQb2+F3EFSdxHx/1WuCmeV/EEs9Bxpt7hPCym7rRddcMChxhEHxuIXsXpo3DBu5cFEDgqttMV8s3klUkbcSq4DhevDA3d9YSFH2cq0GkZYrcXuWfjM5bScVbn5t33p8D4KxWwIT6fdrvsjcZoRxM74JVaBBGoWuSqKMuJcyzPbALK4dUUg/fgKTPaLpNCOpMba4gqDFXiK4DGX2Rj23eVK+JTeNz0Zitz6iE0iFqJfCLKxRZedco4c08pX/VEve2Y+W7HGH5phKmvbexx9XSnPUWcqAF4/yY6LcFDZjZQmoGzWqpzJJbdUKlWFlbRUDbSqfJBryeOe1kb+1FlsAUbsM3gBzCpKVcNTnzw+6yOY4RgRJOR4LaaOD9vIURvN6hN7DcRci1IJPtxTJsHSW0Xp2sgd0GABkZ5yd1fFgKe4hGONiZUdEutqJzOHkFs5/l6eue8ALMUhgpm94IZ/DK9mySwmM6/PKoCgnk3qthhqva7BVGHRr8f+dgjP3Muv8KQ9Kj/SpLiuJrF1MtP8dWPOxXyyKtP5bxdQZtaUaw+ssuWhdPXVH7hrEz1AVzjGJJpoVE5Ba65A6RjKR0RuxivQrIVkJBP9KkqNjZ7eMVM6IxdwyyNRsBf/XzvIyVWK215CkCZHjeAifl1ravcPSUtkqBQU4KnMU3LexqZ+BGx8lMqjCa9wE0ohF1G9Rzf5L5QpVP1E3LewOI3GYHch2nM+vgqU6m7+e9XPpM2btZf4QgKevnyX2pH2G/HhoXdDAH7nk5lsSZN8ISKTSNvCtqO0Lxluz0CzmfkWGfSwwk2/c9BbQ8oSVsd1LlzVcJBgkYQ5JdSeZbTVMKCCawDTw22lSrttA6KNLRV8z0D9V31x0UatTc3cmxFC3liAL/waIQvdE1B++jE+JYGvWHwQ8rbw+Z29AD/9NUYgq7tdajWvqn6EE0CewUcOoLril5kJ/hXCMAmy1DkvUVV6NqjNVc8w80ReUHbNu//Dbn4XyYHvGBfkalWw3lOdPRcnAH1J24NJvjQZaQkZ+j93Bqi+dbSrXrU3I8nQRiK77X0JI6zEGKBOt7Pj1G2prk8TjEDjib1o3/8NPLx45AdVlrzp5vbA7ZL1yCxTnaMmXoEBljKuj8ThmBDJKnixFWi+vAnizcrYQ7VMaU71FxD9VF6pXpuQnIuKy37FfqItiKg1ynvvQzXSqsugFyugOuqi9kwU9Z9C75rfMfnqPuTmmtMunfKtxqYpdXoLYWrjT6bmhBSY2I4Tf3sAqZ3lewJTnJsvhoKuruCwmAsZmlhoKna/HCRCVIhGW1BFLkG3ZUQCtUiHhiButjV75XzXPkv7N9r9ARyLGJn1COXvpxf/Ao35EmzX5FhZoWnHGQ0GUTEecAUFcn2HsG2ROW3+TqcXj1/F94Z4VjsRE79H1POG3dtcI6Ny/9OwYTXh/D41uuIdXOUaCUAZTQdmdMpKdktoWnkud42GhjRGFy6DIm9US5Mr7LbPzMHC8/Mapn8u0v2cshI4vCraW/kSCYQgJsbDu4xJK/g97LqjwA5MuBjCA9DijN8DeJw4S16d32oQJFFMXev6rz2rk5XBVW3Ll+kuelp8I3+veYZGScbXsGYrTrl8OfUVZ4FxPt42pxsgulRpjfU6qqh0XkWVayiG3JjM9vFsVVNYO45872VumoTWDse1orFQppOHwHtiLGCA7SSICPpK5600XIJV2WcVBUud6TO/J8eZZz+wzBdvVThhHd36A8m9qMZoN6vro9oqjJJJMjssJSYKKc9rdjsuf8hWlDMWv5FO1QrRug+5U59t5eVVJuYnnyULGgR7mLmvDxwp9SZhTyWyOuYJO2483/nKcrWhD5m9op/0U6URMpplmd+W0tPDbYbX+cXy48GiIQo43mzboq/XwIH4a6jfB0sSxLtwRea0Ojwk6j9BkADxdaS8l+FnzMwLlM5t18spK53QyPxKwsdyLXqZRmt0QirBAt/P+WHtptbG/+niaj6WcNJQKVJ9oEbZEhRKzbAEYuMePtkzpVpw00DjaBHK/oVhBfciA09dtErEC7Uy3+99S8QYKNsUJ5UTkgpip2W7VxT4H63IWDwK5EH+5dHFxbZVvPf7iAVXRR7dFnaZdW8A3evEakYTGobOKL4MZEjZAZ9SEE+h51KET0Ev7shcRevTB+BIQK3g32TMwjSmR8Juxn/W1oh24v9zRNVYyFJuKldgVmc5bRudVPvLmveCXkYYEno9p7gckfOk5sKnkL/V3jpnoE/NXq9m6vHoGvFCTXKt3ONk2Gp+4Au3tVHgYNyzCUPtnbc68Ig3jdYkJQr7a1axiZrhdVyF0YB2kBvckIkV9/jIWSgDaBlJ0bGJXoKQvWrf12EE1anfLLSRK2wg3HnXl5aalAhQRMnO+kpUfj4J7XOE9Ly0XBCg4AjX9A0MCp/c2rRtfNcoj/a0ggBGF4bqNixnGhWlW9JFKczbnmO76cILh7bk8/4YbBsvoS6dcaphAo9Qg/CB+qtScfdIQ7W3kQl3OOH+s2OlBHlZUL22y50Wm6EFTT4cpXzg0IEyfAXdcielIKqhx7Z2pSh2yUgXgLwkbEENeg+y8R38KiGHE9l9gnqMGOe+Q44f2z4JoAL6Mja1RbOtQxiTTbs7nj+f16fW2VIpDQy1KjsRqHpM6ySfNqhdYhxxuIC/oe8URLhhxbgLBB/MqAlVnJF9HlRIa8xab1Ox4XWD0NzjQ+ZN7Eg7Ex4tihsIhNrKlN+nwvc4Omv2lTEmY/WAJvYyVIwC6A16bmTfCGm7HRhG93SQUQqiqLrZENHkMeEtffWPE8eGjXy/I9Vd7GlfzkP4Zct4dYdxOa8e/mgn9sv/xBRaZvL7lcGaY+H6SiQDkeKWTfs+aOc5TBtjjVy/mN1jn4FPi0a4ts1d/jLbAYQm62RhhFF1teoX6LsC9c5YJHJrfZymzTu9qAUbHQkiriYR4RulXd5HtX82hXDv8Ae56bDTZOrT6P2YNz75yIT81lu5EBHMexUW+fG1fTHG+H4x1W5XZu0krdpBe3QmyJwccvDdJpapQxmh/VayGEAva+6Q3a+60N7B4ovZUfWVTPPy3K9ONMHUVeV8BqzogI9t4vDdyR0cYkHsN2plNdPVLMyPuLAip6aaeEGA507w98at7kaAEHfnUoFHbdRkmux3N7O3WCMu6K+aPxIoO/QGrGYh0Xc6eli3qJ066H3ISnPLUQtQj/WW9VZUcyTYLCYwwnj8Y1fQC/aVqKTXgtz7BV3K2uCfnVq/IWNudne0O8v4gkDWXdlgsvdyZOXO0+NpOS6afrIVbvlM9KLDF5p8qwYKjRXBOv0rOflVmEbsRPFDZa02zMQ982sYQpOq50GjDDTE2ktJr7wRuyQCuxdoRZ7imF/KSqcibViWJdXbTDm0lL1xjTBMwMIkqSuRzATeR3vbYJJCkLABOVVeiXUpk4ENRIel+kaNXHK/LnHeIsNMnOT6TIZgJXq1Sk97Eb7PgcHgVrF5xUh5FaXpATZ3jrp49fLh1YtMnox8mPUlZgUoNuLxcCFsUVgKzAvWdFHayKqDAmW+/s499aU1HyojqJIbpzK1fYruYevDtjrosBI2ibH36KVemOKDIEVPiiDZrLptSFWWuu+d4ChY4kOmh0ZSOGUdHz1zi2mTB7Sdf/AHGywPqoPBHxGFbMbAq0bbnsnL1yyrCoshUS/kZNg3ymQIPdaVwC8vNU38G23+8OoKQfVeH3CwQoVsajhYK8EGYjlet4nU702NaKSKHI2nlK2rKj9wdiyu+odWgdlVdD0m50Zl43bqjBWxKsevv61pOJ1GuasjAJLm98eBdk4agqMIIKa1Oo9bQnQVCkRYGDqsucHWjed0OzHZeIrCjXKpRG5g9NkMLZg9uePwNBkr9SoeqD5zQeKvL3WB2k/i7qDAeiHVOCPlhUvyRZ2fASKtLI4ub2EFpDjpEZ0+sQiQZFhVZijm3FG6mqKBAAgTHK0yKBdT0cQfW8AuRzJ4ga5zL5AUN4Oob33IImdVn3Kc3aVJOydys/CiZkoiRb+iF+Cb44uPfM9QiA74KMUmeDBcGmdygsIEdhkdxeK4m1fPZkh0ofnzD25bJijLa1VKpE65Fej0IOoWeEz6qxyKRxhV39FSTdomaRdp/gPiKB+aSMq9BmbSMF22s8RNNE9AFXYmnLuVmBemv10Crr6CdeqzGPDT1g1ffRMsadEz2FDQoB4rOTfm9XOSAeo0/143oZJ/F1IFPSaYU6di6NIWX6/lfrd1toRrF5lYIeydA8eNXnD9ov5bX4Kr/twNHXjLk6lc5D7dutbruW2sZ4fKFPpYdjYQDKgNLi1atq32GSusv/XSwJi3et5Nh3vwuBWrGaZWcgi/r9178o17TQy8H2kj/9FNixeW2saib+co2IF50zF69dlnS0SwkGSJff+Ndi2tBay/IYesRE4XYzIvymiaJ8raJTiPfK34hzCrAlNMaJYC+cZjB+9f/RhCvgWJ/j7zsemQC24KV5kqI6Va08F4RWMN/psgwaRq1bL5B1YC7quS2BaysUWG9AWynk6FCnQbS/X6ow0NA4O+hygECESCqmX4bt1jWoZT7ZBjieuOd9HL303Jh9c5d2FYFwSI/PVE61GSIkDyOTj+TzG2STPd4iyyBSz3rs5rmxR3Oy0TSI5LHWrP0s7tIT2mcEbcNUnJI5tr/Kd2Z4JLA3MvtGuEc2r9wvMnhHAs8eef3SW0fcHZ904OCl1d2TPKunEpATQg2Hw79CDkhsUmUtQR4ylPo+wk8IdcSfg8Iw7HbWZ5utVNzpfPcTq6zPh5zb37rLdAva81yl6q0/CcKmYfMBUFeGHoXKY+jy8F20iNefBw4KT7VqSRBkV+cfRVPGWueJDWfJy/WAMgyP9PJjil6+PqYbYbOUSdTk/Z/ok5tmxPbm55cHZcQh8hkIZDuudcsDV5ioo/sztsX/vAe0tTxWBUsR3aL1nl1y0/Ask8KIsEMXXNES2eZoNXf5NIStjH8r1EpdK1zIGYzFEoADe9jipiZWBmXNUGmsrm9cpl2DSTF8xYd7VtMDLiYuPkOdCs3iApx2324zUlp5xigzy+RGrgXSnn4lFSrH5MTO3SipO7WxER4TpMDKaJKvTXZ2OoTYp72TZTBQaJ8RV5sqO932jLOOALv+pI9PXmzjKDMVVGSCorRxhu2CpOfMMP8vCd+JacnZgdGfDg5zqg0HVuPsePhnmHcx8Rl9CL5qHwHgeQsQIygDmAEVpCK1gGfsXHc1SUekpWXc6efcfaY8m5U2C291NTUNoEpFbt2kM90sa2mLrS9cAGRem/RRjWXhYB5kwFwtGrk1esZGx8htI1fC4j+SuzPlRFZ4JkT+/PGPURgKxPCLR4uwNDiH4VFuMDQtabxcuno5kGCpC3t2365lVR1sxYQpxh/++VONLEXJmfBtAkPQX/qImXbEuCnXrH+IMMfnV4n+AhgS1enD2NQn7iDEyNuCp9k4Mot66+h13vj9nyHh9fEfP115ZQbzJs9N12mW3zSBDFNuf52ZQeb/YIi5TZDP6lRVJeQX/dm8p7BxXdbrleX4HWjKWVONoYONHi9xtnuzpn+MOZml1mgWa6GAEdmYDxUij/slicAJvCtjWAmsllNBE/l9ZtOfKebYjr6Sbn08uT6oAqoqLjfWkWphk0hZvtKRvBAmHXRhqN3WKW8+1ckYbbXyB7fuCfG+sqxKx2iSqfRMVpjQHMd0Md9PacBUdl66NAo0VLUd6OPgbkSAQU9s19gjjBMzNgg8Fmum5WL7kNx1+4ASS+Lj+9Y5gFJoSVKb5+ZmsgeeSiP3yj0mRYYOYQnm88PZuxyrVtVyoxI6ISBtZuIPK2OUoBWqeXzcgwqRtchssHASEvXz/BpDVIh7DrjnPHQ81z3t3TpFW8HYtEl1L1XKKgrklqoUq9zVXuv52Qi937z1f3MeJUlyV5A+/LSevnWICMZXBRqKOrD178RN3mG8AhCjHGSlNL7miKGMe5EwKcrjnV26wzR+nye7WLTUyijrTTfQITYapv2V6P3TrMm9HpLVf4UwDL8PfY5FeJuQIvdC72+7kIHH0+YRp9h+zqJTH7ADNHx+BIgfYDY/aEt8A3ZbefTUl5sbUP3dkbccyLJxBQ9qaNOH7+JoMy373KJvnoMq3VzF8byvF773PtnvIhjTSUHM0OKGPpva+dsoCHQuYRB2tnQ6ROihHx2Iljqj45ZQeiqGmojLjfyQ4Jp3fiBXkyw7Wfjj5vgUmwaqhsH7F9vit0/fWtzWVURl7B/CS12GSsas1qZuUvIAGMLmx1tSiHr7xipxtY994eC8UEZtL5m7vrG5OM/xaHqXxo8VTfP0TNi+atkEESwMo1uAWfREaWqdY4hXEhk3cWx9zg1jPOqfnJcnlQr29Dmn64/431u55sWsmEaSi0E+m7/PbgXQ4eYeKxSlzCEFKAHfBVuslKgnlrNSAVQyvF4jTdP6ADjzNwEKvUbvvlK5InPw13minVxdr/IGEA9bXQnnKMxkqLoLYXud/TBdksNlCNvKBiIysjQ60nSgFJ8BdYfKPXMMZSlmGkd1S91LSVFYSWGNoY0YqDkdzmrK9KTrPArRdLguuGequf4PadNjuBoxlmLzLTCbO2/Os886nIEhuLFaRhygj6p2NAnXlH7UoiPNuEMrNcBXOXwtM8hcB9uUcrF1z0GJ+L7ceja9x/r+tt2LeC82E2+XJ9e+BnKAln8m1YGU2QxMIHYfblrjLuxQr1cz8JbJCrM4tJRI9ZNbmn/3CQNUf640Hwgi+i3xqyb2DhsjQ0IH5zaR5vHlDseol00M3F8dJhKGp0P3Xe9t6iiYefAPmm7Vmrx8gUbmtxqoO2EkKwZPqfeng33/t/zwbyYuCvWeHyDnUHYe3U4WAiiBPL/1oJd5Fn8lB/9EDjx8URLk09WrIf7s5Sde+1GVoT/y31Km3EXv/YrjqqKReAT+qWpSJnDtw2mcs1Hz2ktjxIGSPpEVaNxbtvTbEZA+mReq2jCToJkE1AW7xb7K8H6PGtOE9R2xVG3hMMxYW14jrwQHX/Nogah7TZxJwMQ2WdvmasjZYbj91c8s8/iv5YOdENt+PUc9MSxqliXHkP79sP27Et6/TKo0shLnsYLacruEYX8WOoupfqEr5um4YKDCFpObdWtcwH5mNgxDI0R4P0YIL3rBZGuTMIGx+K0lHMbXHidVtwojtyWWdCAI9Vua+/UeSHcLEDbr5qMQzR+Z3SUidmdNF+SkZxrj5yJH0nzqVGLjadcifmv+f1MgaPh/aqMvLejaRliP3EapEioKG8nBqGeLzggLUKv41Mt5hGkjnmir+dINoGx9xPqvi2Vbfn/5zI2KxvSL/dfJm7KpjKU43YZAdGROTsUYcPcbJVqdaOREKLiHsxGvtfdu676apgwgSHW7p/jC2eJHH1wfX632WMdG+MpJs30xuyNkm3uqW9DQUpo501Gt5rzO1QB1X5lewBcySY1AprN+LSeo4zThMrZqMbIAczmwNHBsxrwbwzLffSU+2QBrgpxdBdFH3q0s7sXoX/e+vOHld6LrP4sU1CubnDuZTTj6HJaGYj6p0Mzmo+ET9wL0fmiwz8vpepSZLOJGD36RLWMQV1JVMdJI1lrv/CWm4N6VaY0QQv0LiyKb21RLgb3ZEKvofE+ksInEQ4cmsPiXXUGwGaFKp+JhdaalSPn3Z5zmdL0xY4ngPwAw+Y92rQI+ZvWAF3OIgmKl7YfAF9PRQ3c/PywZ0Z35ulMwTAutqexx+oe5EgZ6kkLhm2LPfjhbwQG6PRCnRVEWlZ4HNnvcXmdl9VLPZYFPFB/xxZiojgnAzfI7rdzPlvTb938Lt1O/OXLc4exxTQV6iFTok5ED/paypXtCt8nL3DKfiy0bCn0qDP0Ry9+HEYwCMtPV4qw7BsZ8SMDxZPEXq/GlDpowLwWPYAyMM8b1iaaPDsLrAB7gGF9MZ22QQw4bms1yga1gCJ0VYB3BRxySnBnMcNC9C5OeXTrgpSSTlZje6oTfDXuSgnl5jNa0isoQsa66s9HSJ2q3V1W0ovaQgPRtBpIfZRoqJowg6tjH+sPEJinc3Zo0Y+4umK7Oc3PFZsfTM/S87sVOZi+DKtyuyurKDwnwFwQrrSo5Z5oV+6uXoIvFwvD19J3V/d8s0rf6Y7Sq+9znED2Tv/cAusF/cl9W6PmNU/dU5+WiX/+Vc/mxPOfIpQp5FP7QpAa4xWTwmReaGRxv77KLwTo/y8q7ffmA2fw+pAjJc83JUXxjWU6MEMgkAf5Ijfz8kDI5/gSzWdpxYyz000MNr/AMZM3EQ9eTtG6CyLu5EfHvT7ErJIuyVx4KsI5inN0i8UbYFS3jiuG0JcQoDPQBMvx4D8eSyI4YzHqHR8I50l5qUslOPw6M0+/pWJm27vgXVUcVG4GxPcDBvA/9Ls2xeyMZ20ZJHdeRJ9l3b1QQvdmjq8Y9NE9vJ2BVXO3Afr3zMEu5/fkxoTsgxjkbdGvu+cRIMyqvMiJnMphE1Ct0AO3bZw9VnGEXjasP/yCDphO6QN85WFOrnykNqRTRNU2uqejv/wAK2oCI/g7SDh6atfPlZ/4yM5p8an6qu6bfgTaYfEB7/mKJZJfNEQbY0XbxWbl+7le/e4k+v3BnqQOr3YvBzEHVHcJ6No/P72WGqhEYs5GKuzTLmBAJPp+qaR9I35/yInBDFPgsBtnX1uHMfpDvI0YIjy4rq3+1felnEHNTDDe5Mp3K55zhFeE6SAFq8K7pz+3CxnTR/sHnCYMnGpnTVcoO8SJfRBgT7x1YmI/zpsNZs5vqSWL6RD+9qt5rX+yS0+rTZmUekKra9Uih/tNI5n2wngzQYPtexC7SaKSPwvHb+j9uAvXujS23pfY9Y0/K9LHA+Sow1Dv6qU6kfGA59oEhgZd4JsutrTZBkfygfuc6axgjzkKPeiAIBnHs95ekajPOupQyza9ruihaIdb5251a0BlKqde0iorJTM5YG4YUcb6ZoKOqoNXudvMiNqVTe1LjBmLasr/4UAuXy/LSj8NMA4w+z5hHb8Ra1MYqZbbb2To9ekY/6ePkqtf/kpqLHeVrtdaecUZqEsaSW86IQYC+eVcsvuTwPwY3ntnsngCLDkChOToyuFI1bGkASPnfYHpEPwqoCwXTs4GLFVaJXvjzEY+9L7/SCV2DuEfnkjTUr3/g4QnurCRN0D5TC692Lcxa/ADTD/O6LemndEb9dhnpId9EKZItnw7hDTZ+BOyTjuwY5r7DKLgLAy+pI3C3fB/7RgsQjit0lNgYgfI0d78mxQKPTTZGolWR4zKQFk2b5f+ZikTw5aLcwz0plk2F0KRNTrQmSkfp4AvIJYX5Tpd1edCONfSISvWhIpCTW5tdVDh+8wzRISegFb+8fk9QooyxKQzynqpzL/4dR030d2wnabjHXBzVkBRogthTFlC8LRo9yFzlM08YAKpE6F+K0Ad18Rwi/CwYX3AUjOl9Z8VtQ7qGjn3Hf5u5cBU5y/odEzfOPFzO4uhUWQ8apk2T/rWQQWZGZYNzZPjLefRktJFT0oOCDk/nSwmlkHiLGoye2Z5ZPtMiausFiyBw+jRLv8A6lNiIk9sCLG3KDCsrFAhgLQg+MhO0M2+FBvT6G0I963Jy2FP5ALQJYPlb7+jYXR2TdYVSsbPxs1MGQ9ikYwYUgV7XAKomYuOiqKO9MdPjPEyruYbGZVbQ3p800J+uskT4MbjP3wex7GdKyKdx5KvF1EXWmRoh0bpx0YPKIn2Rw6CiPL2jdHx3EYPRZKtprMRbkgtO86DXVPq4/dVOKT5oHciXBN8fCYYI637sto1NowCtI+DpwGPMKR8Dc6QTzJxb9gu0sxfR6Xv7CXBSJpgGV+0c2lyVToiKa+OqwUDDK0KAF/d4vJtRB+rbinFU82/LxKUwlEgoUgahDKY3gZgvG5yQl858saO6Yz1hQRJJ9rQQmOQ6WvoFGvmOD2a9wKO7iDT28i682RhRvq5wsV7UBMjYkAtv+2Hl9B3wU7gvScAVVKHCBn3kARe39+zRfrIzNU5pnI0veaZ2YxOZ4DBgd+CFs4UvTy+V8p4WEcSnuvs2yuNI4IDj8cjQTVMDz3X94C8YxGvWLUxEan3DURBcVSC6BvcnGHrLjNfTAvgQpQby3ohEr+Fo8CrD9LSFZvTILLzYeVOsFsfr+kMUp3SUJ+uMdkH6YBTyompcflFiIsYc47MY5MQ35EagpCwgPn95TB4S4bYj7V9izjQd+DKfNcqa9sBD4bkowY/qs9jtpG2eR2yc+lGeiCeoCkri19McSg43h5oXhCAuGVvmR7wrlxg2TxBmSw4VWNUrd5gbbQ96AhEU8wxnagiE1hzdzIN9b13AEaCsftg97hGjG12Szs5PYRCP4cUCL/lV61hkkErOyxwoHe+SALiigSF8bt3hZkoAXWeMlw88t0NwYq9UG/JEGLPvFREqFM6bNDEDbZGN4JrOduCrvksSqE1F6q4Z7/pGnF2TaHsjKLhJDAkOnA8e6SQzWr3fjjMoG8m+rkS8iYMdsh9MOCdOAobK0hNXbiYNmyioNV+tHorBwiZQ3eQcXQfq55aKwdxUiPJecSC9gual0M03vpVN8R5eDsIRbRYC4LaaSLRNjetrv8EyfVC/MzuYD6uA2JxaDVNBrfPLffxXoRwRs0uHQqABPqP5ndaG3kPcLOXEF9/x/U6JwSdI9aMS6g5onP9MUjpqTrJcihSeILvOlYvW3ONS1Xt8Yvqk1BXKyI008RUZqdpnw7ujqZWEffjjaIVVGjfdv26JpSraYIocwFLewXwAeok3aqfG7LxXB0Knlze7+uTIPE4+OGnXne4SaOLz+DUSSj5e/8Hfhob2N9ETk1H9K8fiLgY7kCt24EAsL3E+E2ERDZg6RWVwahB7jJyD6v56u2bKiOp7CHDxLiPkvlUwY3VtHh7AviGeP3BlIoLm2I+y1yjO3DTwV2/0PZu5VI7ySsNqtFw9yH65L6o4Tohr136MSLAh/3QkK0wWMv1vYcCDnS8gp6OSnPMEG7o7mlWj0ig1Ml7j+xoiFrQRnXdInV46x/iJWvo2P78omd1hPiRHm/fR5G00isji0fLCy8qcQ1SPb/t9Cc4eerMF2qsaKabL7sbpo6fRvicG0WJGFcO7YOzjnzk2dr93iAKEIcRyzmbYiYgLGAl5zELEUfR44Z6kvsu3+HCsXml23nigJzLAYlTeqaDSNQEg/ORy4JT5JKjmYreHUzmtK8mT26eZOAn8R5M4c3J1eBk/cSq0tQ00bb26eXjfbVYhkm2ilTh4lq/pEFZNi0TY1wNHByEEAW3lph2nmoeB5fiZHkLtdhwMtngHnms7Yb/8xu2RxNZ/YEq+RHbYkx6c1pQ9ca45wl/Hffmy56m2Nmq6KKoacQTB4p2Q4f2OppRkNigf9poI5554+R30AXGlp6xgOQfT9bj5ZyJdJzAylbhjcAedXeCGWfXSilqnpdAheZtnperlVveIVz0aX7h2Rx9HR20AlOQFfo3asNnDLfr8QrduBAkxwK0qQ4tHr4Zml54QrkLr8wvZR6AzRAEyPQ4sT/SgBovFy5i+wUIsrccPXfgbQx9a0uSsR7KP4Gwvs3s7plfSdmC/1t3/iY9Yr5eTF7e12eXrjugR2X0CykUSwGS7Fp0GZNQX72i9WxQ+HEdjDCurzzKFERzN8e3QE114idkazyJ2bxaHGweghv4I4lUaHKoLbsZu9OOOdB++YmH0blkK0KqYJGT/e3+scNxokt6iUWUwaHi1NWcjFk8dD3Y8FJphT0I9sigHTUERnhVTNV8GYA3t7PvRZf0y0heem+VrHx7UOwmuFpWwEus+8tFoQ+JRQdXhztY0IEmJG7oSdq/Wej888/gqDoFeBrnwkYy6NeDRdj2XwwdONo+HKl7d5hhOLKNESvBTMqqwexsCUMtWwpbbn/XbwVXEMpFW8htda0oFE/ZsNL0gL7fg+TzyA3W6tHTjxj0mDC4HYoOtLOjIBElfQwje8J00oKt8d8SLBXwecUAAAA" style="width:32px;height:32px;border-radius:8px;object-fit:cover;object-position:center top;filter:saturate(0.9)"></div>
      <div>
        <div class="ia-header-title">Analista IA</div>
        <div class="ia-header-sub">Analiza toda la base en tiempo real</div>
      </div>
    </div>
    <button class="ia-close" onclick="toggleIA()">&#10005;</button>
  </div>
  <div class="ia-messages" id="ia-messages">
    <div class="ia-msg ia">Hola. Tengo acceso a todos los datos de la base: intervenciones, proveedores, técnicos, activos, repuestos y locales. Puedo analizar, cruzar información, detectar problemas y sugerir acciones. ¿Qué querés revisar?</div>
  </div>
  <div class="ia-suggestions">
    <span class="ia-chip" onclick="enviarIA(\'Analiza todo y dime los 3 problemas mas importantes que ves\')">Análisis general</span>
    <span class="ia-chip" onclick="enviarIA(\'Que proveedor tiene peor desempeño y por que?\')">Proveedores</span>
    <span class="ia-chip" onclick="enviarIA(\'Que activos conviene reemplazar en vez de seguir reparando?\')">Activos críticos</span>
    <span class="ia-chip" onclick="enviarIA(\'Que tecnico interno es mas eficiente y cual tiene mas problemas?\')">Técnicos</span>
    <span class="ia-chip" onclick="enviarIA(\'En que local se gasta mas y por que crees que es asi?\')">Locales</span>
    <span class="ia-chip" onclick="enviarIA(\'Hay repuestos con variacion de precio sospechosa entre proveedores?\')">Repuestos</span>
  </div>
  <div class="ia-input-row">
    <input class="ia-input" id="ia-input" type="text" placeholder="Preguntá o pedile que analice..." onkeydown="if(event.key===\'Enter\') enviarIA()">
    <button class="ia-send" id="ia-send" onclick="enviarIA()">&#8593;</button>
  </div>
</div>

<script>
function toggleIA() {
  const panel = document.getElementById(\'ia-panel\');
  panel.classList.toggle(\'open\');
  if (panel.classList.contains(\'open\')) document.getElementById(\'ia-input\').focus();
}

async function enviarIA(textoDirecto) {
  const input = document.getElementById(\'ia-input\');
  const pregunta = textoDirecto || input.value.trim();
  if (!pregunta) return;
  const msgs = document.getElementById(\'ia-messages\');
  const sendBtn = document.getElementById(\'ia-send\');
  msgs.innerHTML += \'<div class="ia-msg user">\' + pregunta + \'</div>\';
  input.value = \'\';
  sendBtn.disabled = true;
  const thinkId = \'think-\' + Date.now();
  msgs.innerHTML += \'<div class="ia-msg ia thinking" id="\' + thinkId + \'">Analizando la base de datos...</div>\';
  msgs.scrollTop = msgs.scrollHeight;
  try {
    const resp = await fetch(\'/api/ia\', {
      method: \'POST\',
      headers: { \'Content-Type\': \'application/json\' },
      body: JSON.stringify({ pregunta })
    });
    const data = await resp.json();
    document.getElementById(thinkId).remove();
    msgs.innerHTML += \'<div class="ia-msg ia">\' + data.respuesta + \'</div>\';
  } catch(e) {
    document.getElementById(thinkId).remove();
    msgs.innerHTML += \'<div class="ia-msg ia" style="color:var(--danger)">Error al conectar con la IA.</div>\';
  }
  msgs.scrollTop = msgs.scrollHeight;
  sendBtn.disabled = false;
  document.getElementById(\'ia-panel\').classList.add(\'open\');
}
</script>

</body>
</html>'''

@app.route('/')
def portada():
    return PORTADA_HTML

@app.route('/api/empresas')
def api_empresas():
    import glob, os
    folder = os.path.dirname(os.path.abspath(__file__))
    dbs = sorted(glob.glob(os.path.join(folder, '*.db')))
    resultado = []
    for db_path in dbs:
        db_name = os.path.basename(db_path)
        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM intervenciones")
            registros = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM activos WHERE activo=1")
            activos = cur.fetchone()[0]
            conn.close()
            resultado.append({'db': db_name, 'activa': True, 'registros': registros, 'activos': activos})
        except:
            resultado.append({'db': db_name, 'activa': False, 'registros': 0, 'activos': 0})
    # Siempre mostrar al menos 3 slots
    while len(resultado) < 3:
        resultado.append({'db': '', 'activa': False, 'registros': 0, 'activos': 0})
    return jsonify(resultado[:6])

@app.route('/dashboard')
@app.route('/dashboard/<db_name>')
def index(db_name=None):
    global DB_PATH
    if db_name:
        import os
        folder = os.path.dirname(os.path.abspath(__file__))
        DB_PATH = os.path.join(folder, db_name)
    return HTML

# ── API ENDPOINTS ─────────────────────────────────────────────────────────────

def build_where(args):
    conditions = []
    params = []
    if args.get('desde'):
        conditions.append("o.fecha_solicitud >= ?")
        params.append(args['desde'])
    if args.get('hasta'):
        conditions.append("o.fecha_solicitud <= ?")
        params.append(args['hasta'] + ' 23:59:59')
    if args.get('local'):
        conditions.append("l.id = ?")
        params.append(args['local'])
    if args.get('familia'):
        conditions.append("f.id = ?")
        params.append(args['familia'])
    if args.get('tipo'):
        conditions.append("o.tipo_mantenimiento = ?")
        params.append(args['tipo'])
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    return where, params

@app.route('/api/info')
def api_info():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM intervenciones")
    ti = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM activos WHERE activo=1")
    ta = cur.fetchone()[0]
    conn.close()
    return jsonify({'total_intervenciones': ti, 'total_activos': ta, 'db_file': os.path.basename(DB_PATH)})

@app.route('/api/filtros/locales')
def api_filtros_locales():
    conn = get_db()
    data = conn.execute("SELECT id, nombre FROM locales ORDER BY nombre").fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/filtros/familias')
def api_filtros_familias():
    conn = get_db()
    data = conn.execute("SELECT id, nombre FROM familias ORDER BY nombre").fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/resumen/kpis')
def api_resumen_kpis():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT 
            COUNT(DISTINCT i.id) as total_intervenciones,
            ROUND(SUM(c.total), 0) as costo_total,
            ROUND(100.0 * SUM(CASE WHEN o.tipo_mantenimiento='correctivo' THEN 1 ELSE 0 END) / COUNT(*), 1) as pct_correctivo,
            SUM(i.es_reincidencia) as total_reincidencias,
            ROUND(SUM(i.dias_equipo_paralizado), 1) as dias_paralizado,
            COUNT(DISTINCT i.proveedor_id) as proveedores_activos
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id = i.id
        {where}
    """
    row = conn.execute(sql, params).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/resumen/tipos')
def api_resumen_tipos():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT o.tipo_mantenimiento as tipo, COUNT(*) as cantidad
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        {where}
        GROUP BY o.tipo_mantenimiento
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/resumen/familias')
def api_resumen_familias():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT f.nombre as familia, ROUND(SUM(c.total),0) as costo_total
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id = i.id
        {where}
        GROUP BY f.id ORDER BY costo_total DESC
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/resumen/mensual')
def api_resumen_mensual():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT strftime('%Y-%m', o.fecha_solicitud) as mes, COUNT(*) as cantidad
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        {where}
        GROUP BY mes ORDER BY mes
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/resumen/ejecutor')
def api_resumen_ejecutor():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT i.tipo_ejecutor, COUNT(*) as cantidad
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        {where}
        GROUP BY i.tipo_ejecutor
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/resumen/alertas')
def api_resumen_alertas():
    conn = get_db()
    alertas = []
    
    # Proveedor con más reincidencias
    row = conn.execute("""
        SELECT p.nombre, COUNT(*) as reinc FROM intervenciones i
        JOIN proveedores p ON p.id = i.proveedor_id
        WHERE i.es_reincidencia = 1
        GROUP BY p.id ORDER BY reinc DESC LIMIT 1
    """).fetchone()
    if row and row['reinc'] > 5:
        alertas.append({'mensaje': f"Proveedor {row['nombre']} tiene {row['reinc']} reincidencias — revisar calidad del servicio"})

    # Activo con más días paralizado
    row = conn.execute("""
        SELECT a.descripcion, a.codigo, SUM(i.dias_equipo_paralizado) as dias
        FROM intervenciones i JOIN activos a ON a.id = i.activo_id
        GROUP BY a.id ORDER BY dias DESC LIMIT 1
    """).fetchone()
    if row and row['dias'] > 3:
        alertas.append({'mensaje': f"Activo {row['codigo']} — {row['descripcion']} lleva {round(row['dias'],1)} días paralizado acumulados"})

    # % correctivo alto
    row = conn.execute("""
        SELECT ROUND(100.0*SUM(CASE WHEN o.tipo_mantenimiento='correctivo' THEN 1 ELSE 0 END)/COUNT(*),1) as pct
        FROM intervenciones i JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id
    """).fetchone()
    if row and row['pct'] > 60:
        alertas.append({'mensaje': f"El {row['pct']}% de las intervenciones son correctivas — nivel de mantenimiento reactivo alto"})

    conn.close()
    return jsonify(alertas)

@app.route('/api/proveedores')
def api_proveedores():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT p.nombre, p.tipo_contrato,
               COUNT(DISTINCT i.id) as intervenciones,
               ROUND(AVG((julianday(i.fecha_inicio) - julianday(o.fecha_solicitud))*24), 1) as hs_respuesta,
               ROUND(AVG((julianday(i.fecha_fin) - julianday(i.fecha_inicio))*24), 1) as hs_trabajo,
               ROUND(AVG(c.total), 0) as costo_prom,
               ROUND(SUM(c.total), 0) as costo_total,
               SUM(i.es_reincidencia) as reincidencias
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN proveedores p ON p.id = i.proveedor_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id = i.id
        {where}
        AND i.proveedor_id IS NOT NULL
        GROUP BY p.id ORDER BY hs_respuesta
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/tecnicos')
def api_tecnicos():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT t.nombre, ne.nombre as nivel,
               COUNT(DISTINCT i.id) as intervenciones,
               ROUND(AVG((julianday(i.fecha_inicio) - julianday(o.fecha_solicitud))*24), 1) as hs_respuesta,
               ROUND(AVG((julianday(i.fecha_fin) - julianday(i.fecha_inicio))*24), 1) as hs_trabajo,
               SUM(i.es_reincidencia) as reincidencias,
               SUM(i.requirio_segunda_visita) as segunda_visita,
               ROUND(SUM(c.total), 0) as costo_total
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN tecnicos t ON t.id = i.tecnico_id
        JOIN niveles_especializacion ne ON ne.id = t.nivel_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id = i.id
        {where}
        AND i.tecnico_id IS NOT NULL
        GROUP BY t.id ORDER BY hs_trabajo
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/activos')
def api_activos():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT a.codigo, a.descripcion, a.estado, a.costo_reemplazo_estimado as costo_reemplazo,
               l.nombre as local, f.nombre as familia,
               COUNT(DISTINCT i.id) as intervenciones,
               ROUND(SUM(i.dias_equipo_paralizado), 1) as dias_paralizado,
               ROUND(SUM(c.total), 0) as costo_total
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id = i.id
        {where}
        GROUP BY a.id ORDER BY costo_total DESC
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/repuestos')
def api_repuestos():
    conn = get_db()
    where, params = build_where(request.args)
    # Para repuestos el where necesita adaptarse
    sql = f"""
        SELECT r.nombre,
               COUNT(*) as usos,
               ROUND(AVG(ir.precio_unitario), 0) as precio_prom,
               ROUND(MAX(ir.precio_unitario), 0) as precio_max,
               ROUND(MIN(ir.precio_unitario), 0) as precio_min,
               ROUND((MAX(ir.precio_unitario) - MIN(ir.precio_unitario)) / AVG(ir.precio_unitario) * 100, 1) as variacion_pct
        FROM intervencion_repuestos ir
        JOIN repuestos r ON r.id = ir.repuesto_id
        JOIN intervenciones i ON i.id = ir.intervencion_id
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        {where}
        GROUP BY r.id HAVING usos > 0 ORDER BY usos DESC
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/locales')
def api_locales():
    conn = get_db()
    where, params = build_where(request.args)
    sql = f"""
        SELECT l.nombre, c.nombre as ciudad, l.superficie_m2,
               COUNT(DISTINCT i.id) as intervenciones,
               ROUND(100.0 * SUM(CASE WHEN o.tipo_mantenimiento='correctivo' THEN 1 ELSE 0 END) / COUNT(*), 1) as pct_correctivo,
               ROUND(AVG((julianday(i.fecha_inicio) - julianday(o.fecha_solicitud))*24), 1) as hs_respuesta,
               ROUND(SUM(ci.total), 0) as costo_total,
               ROUND(SUM(i.dias_equipo_paralizado), 1) as dias_paralizado,
               CASE WHEN l.superficie_m2 > 0 THEN ROUND(SUM(ci.total)/l.superficie_m2, 0) ELSE NULL END as costo_m2
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id = i.orden_trabajo_id
        JOIN activos a ON a.id = i.activo_id
        JOIN tipos_equipo te ON te.id = a.tipo_equipo_id
        JOIN familias f ON f.id = te.familia_id
        JOIN locales l ON l.id = a.local_id
        JOIN ciudades c ON c.id = l.ciudad_id
        LEFT JOIN costos_intervencion ci ON ci.intervencion_id = i.id
        {where}
        GROUP BY l.id ORDER BY costo_total DESC
    """
    data = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in data])

@app.route('/api/estado-base')
def api_estado_base():
    conn = get_db()
    tablas = ['activos', 'intervenciones', 'proveedores', 'tecnicos', 'locales', 
              'ordenes_trabajo', 'costos_intervencion', 'repuestos', 'intervencion_repuestos']
    resultado = []
    for tabla in tablas:
        try:
            cur = conn.execute(f"SELECT COUNT(*) FROM {tabla}")
            count = cur.fetchone()[0]
            resultado.append({'tabla': tabla, 'registros': count, 'ultima_carga': None})
        except:
            resultado.append({'tabla': tabla, 'registros': 0, 'ultima_carga': None})
    conn.close()
    return jsonify(resultado)


@app.route('/api/cargar', methods=['POST'])
def api_cargar():
    import openpyxl, tempfile, os
    from datetime import datetime

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo'})
    
    archivo = request.files['archivo']
    if not archivo.filename.endswith('.xlsx'):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx'})

    # Guardar temporalmente
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    archivo.save(tmp.name)
    tmp.close()

    resultado = {}

    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        conn = get_db()
        conn.execute("PRAGMA foreign_keys = OFF")

        # ── LOCALES ──────────────────────────────────────────────────────────
        if '5. Locales' in wb.sheetnames:
            ws = wb['5. Locales']
            insertados = errores = 0
            detalles_err = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]: continue
                try:
                    nombre, ciudad, region, unidad, superficie, fecha_ap = (row[i] if i < len(row) else None for i in range(6))
                    # Obtener o crear ciudad
                    cur = conn.execute("SELECT id FROM ciudades WHERE nombre=?", (str(ciudad),))
                    ciudad_row = cur.fetchone()
                    if not ciudad_row:
                        conn.execute("INSERT OR IGNORE INTO regiones(nombre) VALUES(?)", (str(region) if region else 'Sin región',))
                        reg_id = conn.execute("SELECT id FROM regiones WHERE nombre=?", (str(region) if region else 'Sin región',)).fetchone()[0]
                        conn.execute("INSERT INTO ciudades(nombre, region_id) VALUES(?,?)", (str(ciudad), reg_id))
                        ciudad_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    else:
                        ciudad_id = ciudad_row[0]
                    # Obtener o crear unidad de negocio
                    cur = conn.execute("SELECT id FROM unidades_negocio WHERE nombre=?", (str(unidad),))
                    un_row = cur.fetchone()
                    if not un_row:
                        conn.execute("INSERT INTO unidades_negocio(nombre) VALUES(?)", (str(unidad),))
                        un_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    else:
                        un_id = un_row[0]
                    conn.execute("""INSERT OR IGNORE INTO locales(nombre, ciudad_id, unidad_negocio_id, superficie_m2, fecha_apertura)
                        VALUES(?,?,?,?,?)""", (str(nombre), ciudad_id, un_id, superficie, str(fecha_ap) if fecha_ap else None))
                    insertados += 1
                except Exception as e:
                    errores += 1
                    detalles_err.append(f'Fila {row[0]}: {str(e)}')
            resultado['Locales'] = {'insertados': insertados, 'errores': errores, 'detalles': detalles_err[:5]}

        # ── PROVEEDORES ───────────────────────────────────────────────────────
        if '3. Proveedores' in wb.sheetnames:
            ws = wb['3. Proveedores']
            insertados = errores = 0
            detalles_err = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]: continue
                try:
                    nombre, contrato, ciudad, especialidad = (row[i] if i < len(row) else None for i in range(4))
                    activo = row[7] if len(row) > 7 else 1
                    cur = conn.execute("SELECT id FROM ciudades WHERE nombre=?", (str(ciudad) if ciudad else '',))
                    c_row = cur.fetchone()
                    ciudad_id = c_row[0] if c_row else None
                    conn.execute("""INSERT OR IGNORE INTO proveedores(nombre, tipo_contrato, ciudad_id, especialidad, activo)
                        VALUES(?,?,?,?,?)""", (str(nombre), str(contrato) if contrato else 'sin_contrato', ciudad_id, str(especialidad) if especialidad else None, int(activo) if activo else 1))
                    insertados += 1
                except Exception as e:
                    errores += 1
                    detalles_err.append(f'Error: {str(e)}')
            resultado['Proveedores'] = {'insertados': insertados, 'errores': errores, 'detalles': detalles_err[:5]}

        # ── TÉCNICOS ──────────────────────────────────────────────────────────
        if '4. Técnicos' in wb.sheetnames:
            ws = wb['4. Técnicos']
            insertados = errores = 0
            detalles_err = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]: continue
                try:
                    nombre, nivel_nombre, costo_hora, ciudad, fecha_ing, activo = (row[i] if i < len(row) else None for i in range(6))
                    cur = conn.execute("SELECT id FROM niveles_especializacion WHERE nombre=?", (str(nivel_nombre),))
                    niv_row = cur.fetchone()
                    if not niv_row:
                        conn.execute("INSERT INTO niveles_especializacion(nombre, costo_hora) VALUES(?,?)", (str(nivel_nombre), float(costo_hora) if costo_hora else 0))
                        niv_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    else:
                        niv_id = niv_row[0]
                    cur = conn.execute("SELECT id FROM ciudades WHERE nombre=?", (str(ciudad) if ciudad else '',))
                    c_row = cur.fetchone()
                    ciudad_id = c_row[0] if c_row else None
                    conn.execute("""INSERT OR IGNORE INTO tecnicos(nombre, nivel_id, ciudad_base_id, fecha_ingreso, activo)
                        VALUES(?,?,?,?,?)""", (str(nombre), niv_id, ciudad_id, str(fecha_ing) if fecha_ing else None, int(activo) if activo else 1))
                    insertados += 1
                except Exception as e:
                    errores += 1
                    detalles_err.append(f'Error: {str(e)}')
            resultado['Técnicos'] = {'insertados': insertados, 'errores': errores, 'detalles': detalles_err[:5]}

        # ── ACTIVOS ───────────────────────────────────────────────────────────
        if '1. Activos' in wb.sheetnames:
            ws = wb['1. Activos']
            insertados = errores = 0
            detalles_err = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]: continue
                try:
                    codigo, desc, familia_nombre, tipo_nombre, local_nombre, marca, modelo, serie, fecha_inst, vida_util, estado, criticidad, costo_reemplazo, ruta = (row[i] if i < len(row) else None for i in range(14))
                    # Familia
                    cur = conn.execute("SELECT id FROM familias WHERE nombre=?", (str(familia_nombre),))
                    fam_row = cur.fetchone()
                    if not fam_row:
                        conn.execute("INSERT INTO familias(nombre) VALUES(?)", (str(familia_nombre),))
                        fam_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    else:
                        fam_id = fam_row[0]
                    # Tipo equipo
                    cur = conn.execute("SELECT id FROM tipos_equipo WHERE nombre=?", (str(tipo_nombre),))
                    te_row = cur.fetchone()
                    if not te_row:
                        conn.execute("INSERT INTO tipos_equipo(nombre, familia_id, frecuencia_preventivo_dias) VALUES(?,?,?)", (str(tipo_nombre), fam_id, 90))
                        te_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    else:
                        te_id = te_row[0]
                    # Local
                    cur = conn.execute("SELECT id FROM locales WHERE nombre=?", (str(local_nombre),))
                    loc_row = cur.fetchone()
                    local_id = loc_row[0] if loc_row else None
                    conn.execute("""INSERT OR IGNORE INTO activos(codigo, descripcion, tipo_equipo_id, local_id, marca, modelo, numero_serie, fecha_instalacion, vida_util_estimada_anos, estado, criticidad, costo_reemplazo_estimado, activo)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1)""",
                        (str(codigo), str(desc), te_id, local_id, str(marca) if marca else None, str(modelo) if modelo else None,
                         str(serie) if serie else None, str(fecha_inst) if fecha_inst else None,
                         int(vida_util) if vida_util else None, str(estado) if estado else 'bueno',
                         str(criticidad) if criticidad else 'media', float(costo_reemplazo) if costo_reemplazo else None))
                    insertados += 1
                except Exception as e:
                    errores += 1
                    detalles_err.append(f'Error: {str(e)}')
            resultado['Activos'] = {'insertados': insertados, 'errores': errores, 'detalles': detalles_err[:5]}

        # ── INTERVENCIONES ────────────────────────────────────────────────────
        if '2. Intervenciones' in wb.sheetnames:
            ws = wb['2. Intervenciones']
            insertados = errores = 0
            detalles_err = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]: continue
                try:
                    codigo_activo, tipo_mant, fecha_sol, fecha_ini, fecha_fin, tipo_ejec, nombre_ejec, cat_falla, desc_trabajo, resultado_interv, es_reinc, seg_visita, dias_par, mano_obra, rep_total, total_costo, tipo_doc = (row[i] if i < len(row) else None for i in range(17))
                    # Buscar activo
                    cur = conn.execute("SELECT id FROM activos WHERE codigo=?", (str(codigo_activo),))
                    act_row = cur.fetchone()
                    if not act_row: errores += 1; continue
                    activo_id = act_row[0]
                    # Buscar técnico o proveedor
                    tec_id = prov_id = None
                    if tipo_ejec == 'interno':
                        cur = conn.execute("SELECT id FROM tecnicos WHERE nombre=?", (str(nombre_ejec),))
                        t = cur.fetchone()
                        if t: tec_id = t[0]
                    else:
                        cur = conn.execute("SELECT id FROM proveedores WHERE nombre=?", (str(nombre_ejec),))
                        p = cur.fetchone()
                        if p: prov_id = p[0]
                    # Crear orden de trabajo
                    import time
                    num_ot = f"OT-IMP-{int(time.time()*1000)}-{insertados+1}"
                    conn.execute("""INSERT INTO ordenes_trabajo(numero, fecha_solicitud, activo_id, tipo_mantenimiento, estado)
                        VALUES(?,?,?,?,?)""",
                        (num_ot, str(fecha_sol), activo_id, str(tipo_mant) if tipo_mant else 'correctivo', 'cerrada'))
                    ot_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    # Crear intervención
                    conn.execute("""INSERT INTO intervenciones(orden_trabajo_id, activo_id, tipo_ejecutor, tecnico_id, proveedor_id, fecha_inicio, fecha_fin, descripcion_trabajo, resultado, requirio_segunda_visita, es_reincidencia, dias_equipo_paralizado)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (ot_id, activo_id, str(tipo_ejec) if tipo_ejec else 'externo', tec_id, prov_id,
                         str(fecha_ini) if fecha_ini else str(fecha_sol), str(fecha_fin) if fecha_fin else str(fecha_sol),
                         str(desc_trabajo) if desc_trabajo else None, str(resultado_interv) if resultado_interv else 'resuelto',
                         int(seg_visita) if seg_visita else 0, int(es_reinc) if es_reinc else 0,
                         float(dias_par) if dias_par else 0))
                    interv_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    # Crear costo
                    total = float(total_costo) if total_costo else (float(mano_obra or 0) + float(rep_total or 0))
                    conn.execute("""INSERT INTO costos_intervencion(intervencion_id, mano_obra, repuestos_total, total, tipo_documento, tipo_compra)
                        VALUES(?,?,?,?,?,?)""",
                        (interv_id, float(mano_obra) if mano_obra else 0, float(rep_total) if rep_total else 0,
                         total, str(tipo_doc) if tipo_doc else 'factura', 'presupuestada'))
                    insertados += 1
                except Exception as e:
                    errores += 1
                    detalles_err.append(f'Activo {str(row[0]) if row else "?"}: {str(e)}')
            resultado['Intervenciones'] = {'insertados': insertados, 'errores': errores, 'detalles': detalles_err[:5]}

        conn.commit()
        conn.execute("PRAGMA foreign_keys = ON")
        conn.close()

    except Exception as e:
        return jsonify({'error': str(e)})
    finally:
        os.unlink(tmp.name)

    return jsonify({'resultado': resultado})


@app.route('/api/analizar-proformas', methods=['POST'])
def api_analizar_proformas():
    import tempfile, os
    from collections import defaultdict
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se recibio archivo'})
    archivo = request.files['archivo']
    fname = archivo.filename.lower()
    ext = '.xlsx' if 'xlsx' in fname else '.csv'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
    archivo.save(tmp.name)
    tmp.close()
    try:
        rows = []
        headers = []
        if ext == '.csv':
            with open(tmp.name, 'r', encoding='latin-1') as f2:
                content = f2.read()
            sep = ';' if content.count(';') > content.count(',') else ','
            lines = [l for l in content.strip().split('\n') if l.strip()]
            headers = [h.strip() for h in lines[0].split(sep)]
            for line in lines[1:]:
                parts = line.split(sep)
                if len(parts) >= 3:
                    rows.append({headers[i]: parts[i].strip() if i < len(parts) else '' for i in range(len(headers))})
        else:
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name, data_only=True)
            ws = wb.active
            hrow = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h).strip() if h else '' for h in hrow]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v for v in row):
                    rows.append({headers[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else '' for i in range(len(headers))})
        def find_col(names):
            for name in names:
                for h in headers:
                    if name.upper() in h.upper(): return h
            return None
        col_prov = find_col(['PROVEEDOR'])
        col_local = find_col(['LOCAL','SUCURSAL'])
        col_total = find_col(['TOTAL','MONTO'])
        col_fact = find_col(['FACTURA','FACT'])
        col_resp = find_col(['INGENIERO','RESPONSABLE','USUARIO'])
        col_desc = find_col(['DESCRIPCION','DETALLE'])
        def parse_m(val):
            if not val: return 0
            v = str(val).replace('$','').replace(' ','').strip()
            if ',' in v and '.' in v: v = v.replace('.','').replace(',','.')
            elif ',' in v: v = v.replace(',','.')
            try: return float(v)
            except: return 0
        total = len(rows)
        monto_total = 0
        sin_factura = 0
        prov_d = defaultdict(lambda:{'n':0,'m':0,'sf':0})
        loc_d = defaultdict(lambda:{'n':0,'m':0})
        resp_d = defaultdict(lambda:{'n':0,'m':0})
        trab_d = defaultdict(lambda:{'n':0,'m':0})
        KW = ['COMPRESOR','FILTRO','CONDENSADOR','EVAPORADOR','VARIADOR','TABLERO','BOMBA','MOTOR','INSTALACION','MANTENIMIENTO','REPARACION','LIMPIEZA','RECARGA']
        for row in rows:
            p = (row.get(col_prov,'') if col_prov else '').strip().upper() or 'SIN DATO'
            l = (row.get(col_local,'') if col_local else '').strip().upper() or 'SIN DATO'
            r2 = (row.get(col_resp,'') if col_resp else '').strip().upper() or 'SIN DATO'
            desc = (row.get(col_desc,'') if col_desc else '').strip().upper()
            fact = (row.get(col_fact,'') if col_fact else '').strip()
            m = parse_m(row.get(col_total,'') if col_total else '')
            monto_total += m
            if not fact:
                sin_factura += 1
                prov_d[p]['sf'] += 1
            prov_d[p]['n'] += 1; prov_d[p]['m'] += m
            loc_d[l]['n'] += 1; loc_d[l]['m'] += m
            resp_d[r2]['n'] += 1; resp_d[r2]['m'] += m
            for kw in KW:
                if kw in desc:
                    trab_d[kw]['n'] += 1; trab_d[kw]['m'] += m; break
        return jsonify({
            'total': total,
            'monto_total': round(monto_total,0),
            'sin_factura': sin_factura,
            'ticket_prom': round(monto_total/total,0) if total>0 else 0,
            'n_proveedores': len(prov_d),
            'n_locales': len(loc_d),
            'proveedores': sorted([{'p':p,'n':d['n'],'m':round(d['m'],0),'tp':round(d['m']/d['n'],0) if d['n']>0 else 0} for p,d in prov_d.items()],key=lambda x:-x['m'])[:15],
            'sin_factura_list': sorted([{'p':p,'sf':d['sf'],'t':d['n'],'pct':round(d['sf']/d['n']*100,1) if d['n']>0 else 0} for p,d in prov_d.items() if d['sf']>0],key=lambda x:-x['sf'])[:10],
            'locales': sorted([{'l':l,'n':d['n'],'m':round(d['m'],0),'tp':round(d['m']/d['n'],0) if d['n']>0 else 0} for l,d in loc_d.items()],key=lambda x:-x['m'])[:15],
            'ingenieros': sorted([{'r':r2,'n':d['n'],'m':round(d['m'],0),'pct':round(d['m']/monto_total*100,1) if monto_total>0 else 0} for r2,d in resp_d.items()],key=lambda x:-x['m']),
            'trabajos': sorted([{'t':t,'n':d['n'],'m':round(d['m'],0),'tp':round(d['m']/d['n'],0) if d['n']>0 else 0} for t,d in trab_d.items() if d['n']>0],key=lambda x:-x['n'])
        })
    except Exception as e:
        return jsonify({'error': str(e)})
    finally:
        os.unlink(tmp.name)


@app.route('/api/limpiar-carga', methods=['POST'])
def api_limpiar_carga():
    conn = get_db()
    try:
        # Eliminar solo los datos cargados manualmente (prefijo OT-IMP-)
        conn.execute("DELETE FROM intervencion_repuestos WHERE intervencion_id IN (SELECT i.id FROM intervenciones i JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id WHERE o.numero LIKE 'OT-IMP-%')")
        conn.execute("DELETE FROM costos_intervencion WHERE intervencion_id IN (SELECT i.id FROM intervenciones i JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id WHERE o.numero LIKE 'OT-IMP-%')")
        conn.execute("DELETE FROM intervenciones WHERE orden_trabajo_id IN (SELECT id FROM ordenes_trabajo WHERE numero LIKE 'OT-IMP-%')")
        conn.execute("DELETE FROM ordenes_trabajo WHERE numero LIKE 'OT-IMP-%'")
        # Eliminar activos cargados manualmente (prefijo ACT-P o ACT-C)
        conn.execute("DELETE FROM activos WHERE codigo LIKE 'ACT-P%' OR codigo LIKE 'ACT-C%'")
        conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        conn.close()
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/ia', methods=['POST'])
def api_ia():
    data = request.json
    pregunta = data.get('pregunta', '')
    if not pregunta:
        return jsonify({'respuesta': 'Por favor escribi una pregunta.'})

    conn = get_db()

    resumen = dict(conn.execute("""
        SELECT COUNT(DISTINCT i.id) as total_intervenciones,
               ROUND(SUM(c.total),0) as costo_total,
               ROUND(100.0*SUM(CASE WHEN o.tipo_mantenimiento='correctivo' THEN 1 ELSE 0 END)/COUNT(*),1) as pct_correctivo,
               SUM(i.es_reincidencia) as reincidencias,
               ROUND(SUM(i.dias_equipo_paralizado),1) as dias_paralizado
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id=i.id
    """).fetchone())

    proveedores = [dict(r) for r in conn.execute("""
        SELECT p.nombre, p.tipo_contrato,
               COUNT(*) as intervenciones,
               ROUND(AVG((julianday(i.fecha_inicio)-julianday(o.fecha_solicitud))*24),1) as hs_respuesta,
               ROUND(AVG((julianday(i.fecha_fin)-julianday(i.fecha_inicio))*24),1) as hs_trabajo,
               ROUND(AVG(c.total),0) as costo_prom,
               ROUND(SUM(c.total),0) as costo_total,
               SUM(i.es_reincidencia) as reincidencias
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id
        JOIN proveedores p ON p.id=i.proveedor_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id=i.id
        WHERE i.proveedor_id IS NOT NULL
        GROUP BY p.id ORDER BY hs_respuesta
    """).fetchall()]

    tecnicos = [dict(r) for r in conn.execute("""
        SELECT t.nombre, ne.nombre as nivel,
               COUNT(*) as intervenciones,
               ROUND(AVG((julianday(i.fecha_inicio)-julianday(o.fecha_solicitud))*24),1) as hs_respuesta,
               ROUND(AVG((julianday(i.fecha_fin)-julianday(i.fecha_inicio))*24),1) as hs_trabajo,
               SUM(i.es_reincidencia) as reincidencias,
               SUM(i.requirio_segunda_visita) as segunda_visita,
               ROUND(SUM(c.total),0) as costo_total
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id
        JOIN tecnicos t ON t.id=i.tecnico_id
        JOIN niveles_especializacion ne ON ne.id=t.nivel_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id=i.id
        WHERE i.tecnico_id IS NOT NULL
        GROUP BY t.id ORDER BY hs_trabajo
    """).fetchall()]

    activos_top = [dict(r) for r in conn.execute("""
        SELECT a.codigo, a.descripcion, a.estado,
               COALESCE(a.costo_reemplazo_estimado,0) as costo_reemplazo,
               l.nombre as local, f.nombre as familia,
               COUNT(*) as intervenciones,
               ROUND(SUM(i.dias_equipo_paralizado),1) as dias_paralizado,
               ROUND(SUM(c.total),0) as costo_total
        FROM intervenciones i
        JOIN activos a ON a.id=i.activo_id
        JOIN tipos_equipo te ON te.id=a.tipo_equipo_id
        JOIN familias f ON f.id=te.familia_id
        JOIN locales l ON l.id=a.local_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id=i.id
        GROUP BY a.id ORDER BY costo_total DESC LIMIT 10
    """).fetchall()]

    locales = [dict(r) for r in conn.execute("""
        SELECT l.nombre, ci.nombre as ciudad,
               COUNT(*) as intervenciones,
               ROUND(100.0*SUM(CASE WHEN o.tipo_mantenimiento='correctivo' THEN 1 ELSE 0 END)/COUNT(*),1) as pct_correctivo,
               ROUND(AVG((julianday(i.fecha_inicio)-julianday(o.fecha_solicitud))*24),1) as hs_respuesta,
               ROUND(SUM(c.total),0) as costo_total,
               ROUND(SUM(i.dias_equipo_paralizado),1) as dias_paralizado
        FROM intervenciones i
        JOIN ordenes_trabajo o ON o.id=i.orden_trabajo_id
        JOIN activos a ON a.id=i.activo_id
        JOIN locales l ON l.id=a.local_id
        JOIN ciudades ci ON ci.id=l.ciudad_id
        LEFT JOIN costos_intervencion c ON c.intervencion_id=i.id
        GROUP BY l.id ORDER BY costo_total DESC
    """).fetchall()]

    repuestos = [dict(r) for r in conn.execute("""
        SELECT r.nombre, COUNT(*) as usos,
               ROUND(AVG(ir.precio_unitario),0) as precio_prom,
               ROUND(MAX(ir.precio_unitario),0) as precio_max,
               ROUND(MIN(ir.precio_unitario),0) as precio_min,
               ROUND((MAX(ir.precio_unitario)-MIN(ir.precio_unitario))/AVG(ir.precio_unitario)*100,1) as variacion_pct
        FROM intervencion_repuestos ir
        JOIN repuestos r ON r.id=ir.repuesto_id
        GROUP BY r.id HAVING usos > 0 ORDER BY usos DESC
    """).fetchall()]

    conn.close()

    nl = chr(10)
    ctx = f"""Sos un analista experto en gestion de mantenimiento de activos fisicos. Tenes acceso a los datos reales del sistema. Responde en espanol, de forma directa, concreta y util. Usa los numeros para fundamentar. Senala problemas y sugeri acciones especificas.

RESUMEN GENERAL:
- Intervenciones totales: {resumen['total_intervenciones']}
- Costo total: ${resumen['costo_total']:,.0f}
- % correctivo: {resumen['pct_correctivo']}%
- Reincidencias: {resumen['reincidencias']}
- Dias equipos paralizados: {resumen['dias_paralizado']}

PROVEEDORES (por tiempo de respuesta):
{nl.join([f"- {p['nombre']} ({p['tipo_contrato']}): {p['intervenciones']} interv | {p['hs_respuesta']}hs respuesta | {p['hs_trabajo']}hs trabajo | ${p['costo_prom']:,.0f} prom | ${p['costo_total']:,.0f} total | {p['reincidencias']} reincidencias" for p in proveedores])}

TECNICOS INTERNOS (por eficiencia):
{nl.join([f"- {t['nombre']} ({t['nivel']}): {t['intervenciones']} interv | {t['hs_respuesta']}hs respuesta | {t['hs_trabajo']}hs trabajo | {t['reincidencias']} reincidencias | {t['segunda_visita']} 2da visita | ${t['costo_total']:,.0f} total" for t in tecnicos])}

TOP 10 ACTIVOS POR COSTO ACUMULADO:
{nl.join([f"- {a['codigo']} {a['descripcion']} ({a['local']}, {a['familia']}): {a['intervenciones']} interv | {a['dias_paralizado']} dias paralizado | ${a['costo_total']:,.0f} costo acum | ${a['costo_reemplazo']:,.0f} reemplazo | estado: {a['estado']}" for a in activos_top])}

LOCALES (por costo total):
{nl.join([f"- {l['nombre']} ({l['ciudad']}): {l['intervenciones']} interv | {l['pct_correctivo']}% correctivo | {l['hs_respuesta']}hs respuesta | ${l['costo_total']:,.0f} costo | {l['dias_paralizado']} dias paralizado" for l in locales])}

REPUESTOS MAS USADOS:
{nl.join([f"- {r['nombre']}: {r['usos']} usos | prom ${r['precio_prom']:,.0f} | max ${r['precio_max']:,.0f} | min ${r['precio_min']:,.0f} | variacion {r['variacion_pct']}%" for r in repuestos])}
"""

    import urllib.request as urlreq
    payload = json.dumps({
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 1000,
        "system": ctx,
        "messages": [{"role": "user", "content": pregunta}]
    }).encode('utf-8')

    req_obj = urlreq.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={"Content-Type": "application/json", "anthropic-version": "2023-06-01", "x-api-key": os.environ.get("ANTHROPIC_API_KEY", "")}
    )

    try:
        with urlreq.urlopen(req_obj, timeout=30) as resp:
            result = json.loads(resp.read().decode('utf-8'))
            respuesta = result['content'][0]['text']
    except Exception as e:
        respuesta = f"Error al conectar con la IA: {str(e)}"

    return jsonify({'respuesta': respuesta})


if __name__ == '__main__':
    print(f"Iniciando app — base: {DB_PATH}")
    print("Abri http://localhost:5000 en el navegador")
    app.run(debug=False, port=5000)
